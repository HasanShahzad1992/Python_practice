# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# # wd is module, chrome is function,executable path is a parameter, parameter is that thing which is sent to the function,driver path is value of parameter
# website_url="https://en.wikipedia.org/wiki/Pakistan"
# driver.get(website_url)
# target_element=driver.find_elements(By.TAG_NAME,"a")
#  # element is when the tag starts <> and ends </> and the body of text between them
# #  i is just anchor tag
# for i in target_element:
#     link=i.get_attribute("href")
#     print(link)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# #  By is a module which we get from by folder
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# # chromedriver.exe is file name
# # C:/Users/Admin/Desktop/pythonpractice is location name
# # we need to give file name as there a lot of files in the location
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_element=driver.find_elements(By.TAG_NAME,"a")
# for i in target_element:
#     print(i.text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
#
# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By


# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_elements=driver.find_elements(By.CLASS_NAME,"ml-2")
# count=0
# for i in target_elements:
#     count=count+1
#     print(count,"",i.get_attribute("href"))
# driver.quit()

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# for i in range(1,11):
#     path="/html/body/main/div[2]/div/div[1]/table[1]/tbody/tr[" +str(i)+"]/td[3]"
#     target_element=driver.find_element(By.XPATH,path)
#     print(target_element.text)
# driver.quit()
# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# driver_path="C:/Users/AdminDesktop\pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_elements=driver.find_elements(By.TAG_NAME,"a")
# for i in target_elements:
#     print(i.text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
#
# driver_path="C:/Users/AdminDesktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_elements=driver.find_elements(By.CLASS_NAME,"ml-2")
# for i in target_elements:
#     print(i.get_attribute("href"))
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.worldometers.info/coronavirus/")
# wait=WebDriverWait(driver,10)
# target_element=wait.until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[3]/div[3]/div/div[7]/div[1]/div/table/tbody[1]/tr[1]/td[3]")))
# print(target_element.text)
# driver.quit()

# for i in range(1,11):
#     path="/html/body/main/div[2]/div/div[1]/table[1]/tbody/tr["+ str(i) +"]/td[3]"
#     target_element=driver.find_element(By.XPATH,path)
#     print(target_element.text)

# driver.quit()


# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# target_element=wait.until(EC.visibility_of_element_located((By.XPATH,"/html/body/main/div[2]/div/div[1]/table[1]/tbody/tr[2]/td[3]")))
# print(target_element.text)
# driver.quit()

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# for i in range(1,11):
#     path="/html/body/main/div[2]/div/div[1]/table[1]/tbody/tr["+str(i)+"]/td[3]"
#     target_element=driver.find_element(By.XPATH,path)
#     print(target_element.text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# target_element=wait.until(EC.visibility_of_element_located((By.XPATH,"/html/body/main/div[2]/div/div[1]/table[1]/tbody/tr[2]/td[3]")))
# print(target_element.text)
# driver.quit()

import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_elements=driver.find_elements(By.TAG_NAME,"a")
# for i in target_elements:
#     print(i.text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_elements=driver.find_elements(By.CLASS_NAME,"ml-2")
# for i in target_elements:
#     print(i.get_attribute("href"))
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.worldometers.info/coronavirus/")
# target_elements=driver.find_elements(By.CLASS_NAME,"sorting_1")
# for i in target_elements:
#     print(i.text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.worldometers.info/coronavirus/")
# wait=WebDriverWait(driver,10)
# target_element=wait.until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[3]/div[3]/div/div[7]/div[1]/div/table/tbody[1]/tr[9]/td[14]")))
# print(target_element.text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.tppcrpg.net/login.php")
# target_elements=driver.find_elements(By.TAG_NAME,"a")
# for i in target_elements:
#     print(i.text)
# driver.close()

import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_elements=driver.find_elements(By.CLASS_NAME,"mzr-home-logo")
# for i in target_elements:
#     print(i.get_attribute("tabindex"))
# driver.quit()

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_elements=driver.find_elements(By.CLASS_NAME,"mzr-mobile-nav-toggle")
# for i in target_elements:
#     print(i.get_attribute("aria-controls"))
# driver.quit()

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# target_elements=driver.find_elements(By.CLASS_NAME,"mzr-nav-icon mzr-icon-menu")
# for i in target_elements:
#     print(i.get_attribute("xmlns"))
# driver.quit()
# driver_path = "C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver = wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# target_element = wait.until(EC.visibility_of_element_located((By.XPATH,"/html/body/div[3]/nav/a[2]/svg[1]")))
# print(target_element.get_attribute("xmlns"))
# driver.quit()
#
# /html/body/div[3]/nav/a[2]/svg[1]
# import selenium.webdriver as ed
# #
# # import selenium.webdriver as wd
# # from selenium.webdriver.common.by import By
# # from selenium.webdriver.support.ui import WebDriverWait
# # from selenium.webdriver.support import expected_conditions as EC
# # driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# # driver=wd.Chrome(executable_path=driver_path)
# # driver.get("https://moz.com/top500")
# # wait=WebDriverWait(driver,10)
# # target_element=wait.until(EC.visibility_of_element_located((By.XPATH,"/html/body/main/div[2]/div/div[1]/table[1]/thead/tr/th[1]"))
#
# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# for i in range(1,11):
#     path = "/html/body/main/div[2]/div/div[1]/table[1]/tbody/tr[" + str(i)+"]/td[3]"
#     target_element=wait.until(EC.visibility_of_element_located((By.XPATH,path)))
#     print(i,"",target_element.text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://statisticstimes.com/demographics/country/pakistan-cities-population.php")
# wait=WebDriverWait(driver,0)
#  target_elements=wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR,"table#table_id tbody tr")))
# # #tr=[list_1,list_2,list_3,List_4]
# # list_1=[td1,td2,td3,td4,h1,h2,p]
# # td=[6,8,9]
# # list_2=[8,7,9,0]
# # for i in tr:
# #     print(i[0]
# for i in target_elements:
#     td=i.find_elements(By.TAG_NAME,"td")
#     print(td[1].text,td[3].text)
# driver.quit()
#
# # for i in target_elements:
# #     print(i[1])

# list1=[1,2,3,4,5]
# list2=[3,4,5,6]
# list3=[4,5,67,4]
#
# list4=[list1,list2,list3]
# for i in list4:
#     print(i[1],i[3])
# td=[1,2,3,4,6]
# print(td[0])

import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://statisticstimes.com/demographics/country/pakistan-cities-population.php")
# target_elements=driver.find_elements(By.TAG_NAME,"a")
# # this above line gives us a list of anchor (a)
# # e.g a=[1,2,3,4,5]
# # so if we use for i in target_elements, it doesnt mean that i contain target_elements,it means i contains list of a
# #  so if we do for i in target_elements:
# # and if we print(i), we would get a list of a, that is [1,2,3,4,5]
# # 1st iteration gives of i is 1, 1 is not i[1]
#
# for i in target_elements:
#     print(i.text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://statisticstimes.com/demographics/country/pakistan-cities-population.php")
# wait=WebDriverWait(driver,10)
# target_elements=wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR,"table#table_id tbody tr")))
# for i in target_elements:
#     list_of_td=i.find_elements(By.TAG_NAME,"td")
#     print(list_of_td[1].text,list_of_td[3].text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://statisticstimes.com/demographics/country/pakistan-cities-population.php")
# wait=WebDriverWait(driver,10)
# target_elements=wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR,"table#table_id tbody tr")))
#
# for i in target_elements:
#     list_of_td=i.find_elements(By.TAG_NAME,"td")
#     print(list_of_td[1].text,list_of_td[3].text)
# driver.quit()


# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://statisticstimes.com/demographics/country/pakistan-cities-population.php")
# wait=WebDriverWait(driver,10)
# target_elements=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table.table-responsive-md tbody tr")))
# for i in target_elements:
#     td_elements=i.find_elements(By.TAG_NAME,"td")
#     print(td_elements[2].text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.worldometers.info/coronavirus/")
# wait=WebDriverWait(driver,10)
# target_elements=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#main_table_countries_today tbody tr:not([class *= ' '])")))
#
#
# for i in target_elements:
#     td_table=i.find_elements(By.TAG_NAME,"td")
#     if len(td_table)==15:
#         country_names = td_table[1].text
#         population = td_table[2].text
#
#
#     country_names=td_table[1].text
#     population=td_table[2].text
#
#
#     print(td_table[1].text,td_table[2].text)

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.worldometers.info/coronavirus/")
# wait=WebDriverWait(driver,10)
# target_elements=wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR,"table#main_table_countries_today tbody tr:not([class *= ' ']")))
# for i in target_elements:

#
# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# target_elements=wait.until(EC.presence_of_all_elements_located((By.TAG_NAME,"table")))
# target_table=target_elements[1]
# target_tr=target_table.find_elements(By.TAG_NAME,"tr")
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     if len(target_td)==4:
#         target_a=target_td[1].find_elements(By.TAG_NAME,"a")
#         print(target_a[1].get_attribute("href"))
#
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# target_elements=wait.until(EC.presence_of_all_elements_located((By.TAG_NAME,"table")))
# target_table=target_elements[3]
# target_tr=target_table.find_elements(By.TAG_NAME,"tr")
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     if len(target_td)==4:
#         target_a=target_td[1].find_elements(By.TAG_NAME,"a")
#         print(target_a[1].get_attribute("href"))
# driver.quit()
#
# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# target_elements=wait.until(EC.presence_of_all_elements_located((By.TAG_NAME,"table")))
# target_table=target_elements[3]
# target_tr=target_table.find_elements(By.TAG_NAME,"tr")
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     if len(target_td)==4:
#         print(target_td[2].text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# target_tables=wait.until(EC.presence_of_all_elements_located((By.TAG_NAME,"table")))
# target_table=target_tables[3]
# target_tr=target_table.find_elements(By.TAG_NAME,"tr")
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     if len(target_td)==4:
#         print(target_td[2].text)
# driver.quit()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://moz.com/top500")
# wait=WebDriverWait(driver,10)
# target_tables=wait.until(EC.presence_of_all_elements_located((By.TAG_NAME,"table")))
# target_table=target_tables[3]
# target_tr=target_table.find_elements(By.TAG_NAME,"tr")
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     if len(target_td)==4:
#         target_a=target_td[1].find_elements(By.TAG_NAME,"a")
#         print(target_a[0].get_attribute("href"))
# driver.quit()
def web_site_to_excel():
    import selenium.webdriver as wd
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import openpyxl as xl
    import os
    import matplotlib.pyplot as plt
    dict_countries_continents={'Andorra': 'EU', 'United Arab Emirates': 'AS', 'Afghanistan': 'AS', 'Antigua and Barbuda': 'NA', 'Anguilla': 'NA', 'Albania': 'EU', 'Armenia': 'AS', 'Angola': 'AF', 'Antarctica': 'AN', 'Argentina': 'SA', 'American Samoa': 'OC', 'Austria': 'EU', 'Australia': 'OC', 'Aruba': 'NA', 'Åland': 'EU', 'Azerbaijan': 'AS', 'Bosnia and Herzegovina': 'EU', 'Barbados': 'NA', 'Bangladesh': 'AS', 'Belgium': 'EU', 'Burkina Faso': 'AF', 'Bulgaria': 'EU', 'Bahrain': 'AS', 'Burundi': 'AF', 'Benin': 'AF', 'Saint Barthélemy': 'NA', 'Bermuda': 'NA', 'Brunei': 'AS', 'Bolivia': 'SA', 'Bonaire, Sint Eustatius, and Saba': 'NA', 'Brazil': 'SA', 'Bahamas': 'NA', 'Bhutan': 'AS', 'Bouvet Island': 'AN', 'Botswana': 'AF', 'Belarus': 'EU', 'Belize': 'NA', 'Canada': 'NA', 'Cocos (Keeling) Islands': 'AS', 'DR Congo': 'AF', 'Central African Republic': 'AF', 'Congo Republic': 'AF', 'Switzerland': 'EU', 'Ivory Coast': 'AF', 'Cook Islands': 'OC', 'Chile': 'SA', 'Cameroon': 'AF', 'China': 'AS', 'Colombia': 'SA', 'Costa Rica': 'NA', 'Cuba': 'NA', 'Cabo Verde': 'AF', 'Curaçao': 'NA', 'Christmas Island': 'OC', 'Cyprus': 'EU', 'Czechia': 'EU', 'Germany': 'EU', 'Djibouti': 'AF', 'Denmark': 'EU', 'Dominica': 'NA', 'Dominican Republic': 'NA', 'Algeria': 'AF', 'Ecuador': 'SA', 'Estonia': 'EU', 'Egypt': 'AF', 'Western Sahara': 'AF', 'Eritrea': 'AF', 'Spain': 'EU', 'Ethiopia': 'AF', 'Finland': 'EU', 'Fiji': 'OC', 'Falkland Islands': 'SA', 'Micronesia': 'OC', 'Faroe Islands': 'EU', 'France': 'EU', 'Gabon': 'AF', 'United Kingdom': 'EU', 'Grenada': 'NA', 'Georgia': 'AS', 'French Guiana': 'SA', 'Guernsey': 'EU', 'Ghana': 'AF', 'Gibraltar': 'EU', 'Greenland': 'NA', 'The Gambia': 'AF', 'Guinea': 'AF', 'Guadeloupe': 'NA', 'Equatorial Guinea': 'AF', 'Greece': 'EU', 'South Georgia and South Sandwich Islands': 'AN', 'Guatemala': 'NA', 'Guam': 'OC', 'Guinea-Bissau': 'AF', 'Guyana': 'SA', 'Hong Kong': 'AS', 'Heard and McDonald Islands': 'AN', 'Honduras': 'NA', 'Croatia': 'EU', 'Haiti': 'NA', 'Hungary': 'EU', 'Indonesia': 'AS', 'Ireland': 'EU', 'Israel': 'AS', 'Isle of Man': 'EU', 'India': 'AS', 'British Indian Ocean Territory': 'AS', 'Iraq': 'AS', 'Iran': 'AS', 'Iceland': 'EU', 'Italy': 'EU', 'Jersey': 'EU', 'Jamaica': 'NA', 'Jordan': 'AS', 'Japan': 'AS', 'Kenya': 'AF', 'Kyrgyzstan': 'AS', 'Cambodia': 'AS', 'Kiribati': 'OC', 'Comoros': 'AF', 'St Kitts and Nevis': 'NA', 'North Korea': 'AS', 'South Korea': 'AS', 'Kuwait': 'AS', 'Cayman Islands': 'NA', 'Kazakhstan': 'AS', 'Laos': 'AS', 'Lebanon': 'AS', 'Saint Lucia': 'NA', 'Liechtenstein': 'EU', 'Sri Lanka': 'AS', 'Liberia': 'AF', 'Lesotho': 'AF', 'Lithuania': 'EU', 'Luxembourg': 'EU', 'Latvia': 'EU', 'Libya': 'AF', 'Morocco': 'AF', 'Monaco': 'EU', 'Moldova': 'EU', 'Montenegro': 'EU', 'Saint Martin': 'NA', 'Madagascar': 'AF', 'Marshall Islands': 'OC', 'North Macedonia': 'EU', 'Mali': 'AF', 'Myanmar': 'AS', 'Mongolia': 'AS', 'Macao': 'AS', 'Northern Mariana Islands': 'OC', 'Martinique': 'NA', 'Mauritania': 'AF', 'Montserrat': 'NA', 'Malta': 'EU', 'Mauritius': 'AF', 'Maldives': 'AS', 'Malawi': 'AF', 'Mexico': 'NA', 'Malaysia': 'AS', 'Mozambique': 'AF', 'Namibia': 'AF', 'New Caledonia': 'OC', 'Niger': 'AF', 'Norfolk Island': 'OC', 'Nigeria': 'AF', 'Nicaragua': 'NA', 'Netherlands': 'EU', 'Norway': 'EU', 'Nepal': 'AS', 'Nauru': 'OC', 'Niue': 'OC', 'New Zealand': 'OC', 'Oman': 'AS', 'Panama': 'NA', 'Peru': 'SA', 'French Polynesia': 'OC', 'Papua New Guinea': 'OC', 'Philippines': 'AS', 'Pakistan': 'AS', 'Poland': 'EU', 'Saint Pierre and Miquelon': 'NA', 'Pitcairn Islands': 'OC', 'Puerto Rico': 'NA', 'Palestine': 'AS', 'Portugal': 'EU', 'Palau': 'OC', 'Paraguay': 'SA', 'Qatar': 'AS', 'Réunion': 'AF', 'Romania': 'EU', 'Serbia': 'EU', 'Russia': 'EU', 'Rwanda': 'AF', 'Saudi Arabia': 'AS', 'Solomon Islands': 'OC', 'Seychelles': 'AF', 'Sudan': 'AF', 'Sweden': 'EU', 'Singapore': 'AS', 'Saint Helena': 'AF', 'Slovenia': 'EU', 'Svalbard and Jan Mayen': 'EU', 'Slovakia': 'EU', 'Sierra Leone': 'AF', 'San Marino': 'EU', 'Senegal': 'AF', 'Somalia': 'AF', 'Suriname': 'SA', 'South Sudan': 'AF', 'São Tomé and Príncipe': 'AF', 'El Salvador': 'NA', 'Sint Maarten': 'NA', 'Syria': 'AS', 'Eswatini': 'AF', 'Turks and Caicos Islands': 'NA', 'Chad': 'AF', 'French Southern Territories': 'AN', 'Togo': 'AF', 'Thailand': 'AS', 'Tajikistan': 'AS', 'Tokelau': 'OC', 'Timor-Leste': 'OC', 'Turkmenistan': 'AS', 'Tunisia': 'AF', 'Tonga': 'OC', 'Turkey': 'AS', 'Trinidad and Tobago': 'NA', 'Tuvalu': 'OC', 'Taiwan': 'AS', 'Tanzania': 'AF', 'Ukraine': 'EU', 'Uganda': 'AF', 'U.S. Outlying Islands': 'OC', 'United States': 'NA', 'Uruguay': 'SA', 'Uzbekistan': 'AS', 'Vatican City': 'EU', 'St Vincent and Grenadines': 'NA', 'Venezuela': 'SA', 'British Virgin Islands': 'NA', 'U.S. Virgin Islands': 'NA', 'Vietnam': 'AS', 'Vanuatu': 'OC', 'Wallis and Futuna': 'OC', 'Samoa': 'OC', 'Kosovo': 'EU', 'Yemen': 'AS', 'Mayotte': 'AF', 'South Africa': 'AF', 'Zambia': 'AF', 'Zimbabwe': 'AF'}

    driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
    driver=wd.Ch
    rome(executable_path=driver_path)
    driver.get("https://tradingeconomics.com/matrix")
    wait=WebDriverWait(driver,10)
    target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
    if os.path.exists("inflation.xlsx"):
        workbook=xl.load_workbook("inflation.xlsx")
    else:
        workbook=xl.Workbook()
    sheetnames=workbook.sheetnames
    sheetname="inflation"
    if sheetname in sheetnames:
        sheet_to_remove=workbook[sheetname]
        workbook.remove(sheet_to_remove)

        #remove function doesnt return so thats why we dont write sheet 1 as variable, wheras .create gives us a return in form of sheet
    sheet1 = workbook.create_sheet(sheetname)





    for i in target_tr:
        target_td=i.find_elements(By.TAG_NAME,"td")
        if len(target_td)>=11:
            target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
            target_gdp_a = target_td[1].find_element(By.TAG_NAME, "a")
            print(target_country_a.text,target_gdp_a.text)
            if target_country_a.text in dict_countries_continents.keys():
                continent_name = dict_countries_continents[target_country_a.text]
                sheet1.append([target_country_a.text, target_gdp_a.text, continent_name])
            else:
                continue



    workbook.save("inflation.xlsx")
            # i am appending not in list, but in sheet1(excel),you can append not only in list but also in excel

    driver.quit()
web_site_to_excel()

# def excel_to_graph():
#     import matplotlib.pyplot as plt
#     import openpyxl as xl
#     workbook=xl.load_workbook("inflation.xlsx")
#     sheet1=workbook["inflation"]
#     continent_dict={}

    # # country_inflation_dict={}
    # # desired_countries=[]
    # # print("please give your top 5 countries that you want to see in your bar chart")
    # # for i in range(5):
    # #     desired_country=input("please enter your country name")
    # #     desired_countries.append(desired_country)
    # for i in sheet1.iter_rows(values_only=True):
    #     country=i[0]
    #     gdp_rate=i[1]
    #     continent=i[2]
    #     if gdp_rate==None or gdp_rate=="":
    #         continue
    #     if continent not in continent_dict:
    #         continent_dict[continent]=int(gdp_rate)
    #     else:
    #         continent_dict[continent]=continent_dict[continent] + int(gdp_rate)
    #     # if country in desired_countries:
    #     #     country_inflation_dict[country]=float(inflation_rate)

web_site_to_excel()






#     print(continent_dict)
#     # plt.bar(country_inflation_dict.keys(),country_inflation_dict.values())
#     plt.bar(continent_dict.keys(),continent_dict.values())
#
#     # plt.xlabel("count")
#     plt.xlabel("continent")
#     plt.ylabel("gdp")
#     # plt.ylabel("inflation rate")
#     plt.title("inflation rates of different countries")
#     plt.show()
# while True:
#     print("do you want to extract data from website to excel so press 1")
#     print("do you want to extract data from website to excel so press 2")
#     print("press 3 to break out of loop")
#     choice=int(input("choose your choice"))
#     if choice == 1:
#         web_site_to_excel()
#     elif choice == 2:
#         excel_to_graph()
#     elif choice == 3:
#         break
#     else:
#         print("wrong choice, please try again")
#
#

