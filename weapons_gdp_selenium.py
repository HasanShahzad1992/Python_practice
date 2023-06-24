import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import matplotlib.pyplot as plt
import os
import openpyxl as xl
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://data.worldbank.org/indicator/MS.MIL.XPND.GD.ZS?locations=AF")
# wait=WebDriverWait(driver,10)
# target_class=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"div.item:not([class *= ' ']")))
# dict_country_gdp_percent={}
# for i in target_class:
#     target_div=i.find_elements(By.TAG_NAME,"div")
#     target_country=target_div[0].text
#     target_gdp_percent = target_div[2].text
#     print(target_country)
#     print(target_gdp_percent)
#     if target_gdp_percent==None or target_gdp_percent=="":
#         continue
#     else:
#         dict_country_gdp_percent[target_country] = target_gdp_percent
# print(dict_country_gdp_percent)
dict_country_gdp_percent={'Afghanistan': '1.4', 'Albania': '1.4', 'Algeria': '5.6', 'Angola': '1.1', 'Argentina': '0.6', 'Armenia': '4.4', 'Australia': '2.0', 'Austria': '0.8', 'Azerbaijan': '5.3', 'Bahrain': '3.6', 'Bangladesh': '1.3', 'Belarus': '1.2', 'Belgium': '1.1', 'Belize': '1.3', 'Benin': '0.6', 'Bolivia': '1.5', 'Bosnia and Herzegovina': '0.9', 'Botswana': '2.9', 'Brazil': '1.2', 'Brunei Darussalam': '3.3', 'Bulgaria': '1.6', 'Burkina Faso': '2.9', 'Burundi': '1.9', 'Cabo Verde': '0.5', 'Cambodia': '2.3', 'Cameroon': '1.0', 'Canada': '1.3', 'Central African Republic': '1.9', 'Chad': '2.8', 'Chile': '2.0', 'China': '1.7', 'Colombia': '3.4', 'Congo, Dem. Rep.': '0.6', 'Congo, Rep.': '2.9', "Cote d'Ivoire": '1.3', 'Croatia': '2.7', 'Cuba': '2.9', 'Cyprus': '2.0', 'Czechia': '1.4', 'Denmark': '1.4', 'Djibouti': '3.7', 'Dominican Republic': '0.7', 'Ecuador': '2.5', 'Egypt, Arab Rep.': '1.3', 'El Salvador': '1.4', 'Equatorial Guinea': '1.3', 'Eritrea': '20.9', 'Estonia': '2.2', 'Eswatini': '1.7', 'Ethiopia': '0.5', 'Fiji': '1.7', 'Finland': '2.0', 'France': '1.9', 'Gabon': '1.7', 'Gambia, The': '0.7', 'Georgia': '1.7', 'Germany': '1.3', 'Ghana': '0.5', 'Greece': '3.9', 'Guatemala': '0.4', 'Guinea': '1.5', 'Guinea-Bissau': '1.8', 'Guyana': '1.0', 'Haiti': '0.2', 'Honduras': '1.6', 'Hungary': '1.6', 'India': '2.7', 'Indonesia': '0.7', 'Iran, Islamic Rep.': '2.3', 'Iraq': '2.3', 'Ireland': '0.3', 'Israel': '5.2', 'Italy': '1.5', 'Jamaica': '1.4', 'Japan': '1.1', 'Jordan': '5.0', 'Kazakhstan': '0.9', 'Kenya': '1.1', 'Korea, Rep.': '2.8', 'Kosovo': '0.9', 'Kuwait': '6.7', 'Kyrgyz Republic': '1.7', 'Lao PDR': '0.2', 'Latvia': '2.3', 'Lebanon': '3.0', 'Lesotho': '1.3', 'Liberia': '1.0', 'Libya': '15.5', 'Lithuania': '2.0', 'Luxembourg': '0.6', 'Madagascar': '0.8', 'Malawi': '0.9', 'Malaysia': '1.1', 'Mali': '2.8', 'Malta': '0.5', 'Mauritania': '2.4', 'Mauritius': '0.2', 'Mexico': '0.7', 'Moldova': '0.4', 'Mongolia': '0.7', 'Montenegro': '1.7', 'Morocco': '4.2', 'Mozambique': '1.5', 'Myanmar': '3.3', 'Namibia': '3.0', 'Nepal': '1.3', 'Netherlands': '1.4', 'New Zealand': '1.4', 'Nicaragua': '0.6', 'Niger': '1.8', 'Nigeria': '1.0', 'North Macedonia': '1.5', 'Norway': '1.8', 'Oman': '7.3', 'Pakistan': '3.8', 'Panama': '1.0', 'Papua New Guinea': '0.4', 'Paraguay': '1.0', 'Peru': '1.1', 'Philippines': '1.0', 'Poland': '2.1', 'Portugal': '2.1', 'Qatar': '4.8', 'Romania': '2.0', 'Russian Federation': '4.1', 'Rwanda': '1.5', 'Saudi Arabia': '6.6', 'Senegal': '1.8', 'Serbia': '2.1', 'Seychelles': '1.5', 'Sierra Leone': '0.6', 'Singapore': '3.0', 'Slovak Republic': '1.7', 'Slovenia': '1.2', 'Somalia': '1.5', 'South Africa': '0.9', 'South Sudan': '2.4', 'Spain': '1.4', 'Sri Lanka': '1.9', 'Sudan': '1.0', 'Sweden': '1.3', 'Switzerland': '0.7', 'Syrian Arab Republic': '4.1', 'Tajikistan': '1.0', 'Tanzania': '1.1', 'Thailand': '1.3', 'Timor-Leste': '1.1', 'Togo': '1.8', 'Trinidad and Tobago': '0.9', 'Tunisia': '2.8', 'Turkiye': '2.1', 'Turkmenistan': '2.9', 'Uganda': '2.4', 'Ukraine': '3.2', 'United Arab Emirates': '5.6', 'United Kingdom': '2.2', 'United States': '3.5', 'Uruguay': '2.3', 'Uzbekistan': '3.6', 'Venezuela, RB': '0.2', 'Vietnam': '2.3', 'Yemen, Rep.': '4.0', 'Zambia': '1.3', 'Zimbabwe': '0.0', 'World': '2.2', 'Arab World': '4.6', 'Caribbean small states': '1.0', 'Central Europe and the Baltics': '1.9', 'East Asia & Pacific': '1.7', 'East Asia & Pacific (excluding high income)': '1.6', 'Euro area': '1.5', 'Europe & Central Asia': '1.8', 'Europe & Central Asia (excluding high income)': '3.1', 'European Union': '1.5', 'Fragile and conflict affected situations': '1.6', 'Heavily indebted poor countries (HIPC)': '1.2', 'Latin America & Caribbean': '1.2', 'Latin America & Caribbean (excluding high income)': '1.1', 'Least developed countries: UN classification': '1.4', 'Middle East & North Africa': '4.4', 'Middle East & North Africa (excluding high income)': '2.7', 'North America': '3.3', 'OECD members': '2.4', 'Other small states': '3.3', 'Pacific island small states': '1.0', 'Small states': '3.1', 'South Asia': '2.6', 'Sub-Saharan Africa': '1.1', 'Sub-Saharan Africa (excluding high income)': '1.1', 'High income': '2.5', 'Low & middle income': '1.8', 'Low income': '1.2', 'Lower middle income': '2.1', 'Middle income': '1.8', 'Upper middle income': '1.8'}

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://tradingeconomics.com/matrix")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
# if os.path.exists("gdp_weapons.xlsx"):
#     workbook=xl.load_workbook("gdp_weapons.xlsx")
# else:
#     workbook=xl.Workbook()
# sheetnames=workbook.sheetnames
# sheetname="weapon_spending"
# if sheetname in sheetnames:
#     worksheet_to_remove=workbook[sheetname]
#     workbook.remove(worksheet_to_remove)
# sheet1=workbook.create_sheet(sheetname)
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
#     target_gdp_a=target_td[1].find_element(By.TAG_NAME,"a")
#     if target_country_a.text in dict_country_gdp_percent:
#         value_gdp_percent=float(dict_country_gdp_percent[target_country_a.text])
#         gdp_actual_value=value_gdp_percent*(int(target_gdp_a.text))
#         sheet1.append([target_country_a.text,gdp_actual_value])
#     else:
#         continue
# workbook.save("gpd_weapons.xlsx")
# driver.quit()

workbook=xl.load_workbook("gpd_weapons.xlsx")
sheet1=workbook["weapon_spending"]
countries_weapons=[]
dict_country_gdp_weapons={}

for i in range(5):
    print("please choose your country")
    countries_choice=input("choose your country")
    countries_weapons.append(countries_choice)
for i in sheet1.iter_rows(values_only=True):
    countries=i[0]
    gdp_spending_weapons=i[1]
    if countries in countries_weapons:
        dict_country_gdp_weapons[countries] = gdp_spending_weapons
    else:
        continue
print(dict_country_gdp_weapons)
plt.bar(list(dict_country_gdp_weapons.keys()),list(dict_country_gdp_weapons.values()))
plt.xlabel("countries")
plt.ylabel("weapons_gdp_spending")
plt.title("top 5 country spending on weapons")
plt.show()


















