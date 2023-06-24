import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import matplotlib.pyplot as plt
import openpyxl as xl
import os

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.numbeo.com/cost-of-living/rankings.jsp")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#t2 tbody tr")))
# if os.path.exists("cost_living.xlsx"):
#     workbook=xl.load_workbook("cost_living.xlsx")
# else:
#     workbook=xl.Workbook()
# sheetnames=workbook.sheetnames
# sheetname="cost_of_living"
# if sheetname in sheetnames:
#     sheet_to_remove=workbook[sheetname]
#     workbook.remove(sheet_to_remove)
# sheet1=workbook.create_sheet(sheetname)
# list=[]
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     target_cities=target_td[1].text
#     specific_cities_countries=target_cities.split(",")
#     specific_city=specific_cities_countries[0].lower()
#     specific_country=specific_cities_countries[len(specific_cities_countries)-1].lower()
#     target_cost_living_index=target_td[2].text
#     sheet1.append([specific_city,specific_country,target_cost_living_index])
#     print(specific_city,specific_country,target_cost_living_index)
#     # target_cost_living=target_td[2].text
#     # print(target_cities,target_cost_living)
#     # list.append([target_cities,target_cost_living])
# workbook.save("cost_of_living.xlsx")
# driver.quit()
print(1)
workbook=xl.load_workbook("cost_of_living.xlsx")
sheet1=workbook["cost_of_living"]
dict_city_cost_living={}
print("please choose a country name")
choice_country=input("choose a country").lower()
print(choice_country)
for i in sheet1.iter_rows(values_only=True):
    cities=i[0]
    # countries=i[1]
    countries=i[1].strip()
    cost_of_living_index=float(i[2])
    print(countries,len(countries))
    if countries==choice_country:
        dict_city_cost_living[cities]=cost_of_living_index
print(dict_city_cost_living)
plt.bar(dict_city_cost_living.keys(),dict_city_cost_living.values())
plt.xlabel("cities")
plt.ylabel("cost of living index")
plt.title("cost of living index for different cities")
plt.show()


