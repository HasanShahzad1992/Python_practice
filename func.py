# # def website_to_excel():
# #     import selenium.webdriver as wd
# #     from selenium.webdriver.common.by import By
# #     from selenium.webdriver.support.ui import WebDriverWait
# #     from selenium.webdriver.support import expected_conditions as EC
# #     import openpyxl as xl
# #     import os
# #
# #     driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# #     driver=wd.Chrome(executable_path=driver_path)
# #     driver.get("https://tradingeconomics.com/matrix")
# #     wait=WebDriverWait(driver,10)
# #     target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
# #
# #     if os.path.exists("inflation.xlsx"):
# #         workbook=xl.load_workbook("inflation.xlsx")
# #     else:
# #         workbook=xl.Workbook()
# #     sheetnames = workbook.sheetnames
# #     sheetname = "inflation"
# #     if sheetname in sheetnames:
# #         sheet_to_remove=workbook[sheetname]
# #         workbook.remove(sheet_to_remove)
# #     sheet1=workbook.create_sheet(sheetname)
# #
# #
# #
# #
# #     for i in target_tr:
# #         target_td=i.find_elements(By.TAG_NAME,"td")
# #         target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
# #         target_inflation_a=target_td[5].find_element(By.TAG_NAME,"a")
# #         print(target_country_a.text,target_inflation_a.text)
# #         sheet1.append([target_country_a.text,target_inflation_a.text])
# #     workbook.save("inflation.xlsx")
# #     driver.quit()
# #
# # def excel_to_graph():
# #     import openpyxl as xl
# #     import matplotlib.pyplot as plt
# #     workbook=xl.load_workbook("inflation.xlsx")
# #     sheet1=workbook["inflation"]
# #     total_countries=[]
# #     inflation_of_total_countries=[]
# #     desired_countries=[]
# #     print("please choose 5 countries for bar")
# #     for i in range(5):
# #
# #         country_i_chose=input("choose your country")
# #         desired_countries.append(country_i_chose)
# #         dict={}
#
#
#     # for i in sheet1.iter_rows(values_only=True):
#     #     country=i[0]
#     #     inflation_one_country=i[1]
#     #     total_countries.append(country)
#     #     inflation_of_total_countries.append(inflation_one_country)
#     #     if country in desired_countries:
#     #         dict[country]=float(inflation_one_country)
#     # plt.bar(dict.keys(),dict.values())
#     # plt.xlabel("countries")
#     # plt.ylabel("inflation")
#     # plt.title("inflation of countries")
#     # plt.show()
#
#
#
#
# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# import openpyxl as xl
# import os
# import matplotlib.pyplot as plt
# country_continent_dict={'Andorra': 'EU', 'United Arab Emirates': 'AS', 'Afghanistan': 'AS',
#                              'Antigua and Barbuda': 'NA', 'Anguilla': 'NA', 'Albania': 'EU', 'Armenia': 'AS',
#                              'Angola': 'AF', 'Antarctica': 'AN', 'Argentina': 'SA', 'American Samoa': 'OC',
#                              'Austria': 'EU', 'Australia': 'OC', 'Aruba': 'NA', 'Åland': 'EU', 'Azerbaijan': 'AS',
#                              'Bosnia and Herzegovina': 'EU', 'Barbados': 'NA', 'Bangladesh': 'AS', 'Belgium': 'EU',
#                              'Burkina Faso': 'AF', 'Bulgaria': 'EU', 'Bahrain': 'AS', 'Burundi': 'AF', 'Benin': 'AF',
#                              'Saint Barthélemy': 'NA', 'Bermuda': 'NA', 'Brunei': 'AS', 'Bolivia': 'SA',
#                              'Bonaire, Sint Eustatius, and Saba': 'NA', 'Brazil': 'SA', 'Bahamas': 'NA', 'Bhutan': 'AS',
#                              'Bouvet Island': 'AN', 'Botswana': 'AF', 'Belarus': 'EU', 'Belize': 'NA', 'Canada': 'NA',
#                              'Cocos (Keeling) Islands': 'AS', 'DR Congo': 'AF', 'Central African Republic': 'AF',
#                              'Congo Republic': 'AF', 'Switzerland': 'EU', 'Ivory Coast': 'AF', 'Cook Islands': 'OC',
#                              'Chile': 'SA', 'Cameroon': 'AF', 'China': 'AS', 'Colombia': 'SA', 'Costa Rica': 'NA',
#                              'Cuba': 'NA', 'Cabo Verde': 'AF', 'Curaçao': 'NA', 'Christmas Island': 'OC',
#                              'Cyprus': 'EU', 'Czechia': 'EU', 'Germany': 'EU', 'Djibouti': 'AF', 'Denmark': 'EU',
#                              'Dominica': 'NA', 'Dominican Republic': 'NA', 'Algeria': 'AF', 'Ecuador': 'SA',
#                              'Estonia': 'EU', 'Egypt': 'AF', 'Western Sahara': 'AF', 'Eritrea': 'AF', 'Spain': 'EU',
#                              'Ethiopia': 'AF', 'Finland': 'EU', 'Fiji': 'OC', 'Falkland Islands': 'SA',
#                              'Micronesia': 'OC', 'Faroe Islands': 'EU', 'France': 'EU', 'Gabon': 'AF',
#                              'United Kingdom': 'EU', 'Grenada': 'NA', 'Georgia': 'AS', 'French Guiana': 'SA',
#                              'Guernsey': 'EU', 'Ghana': 'AF', 'Gibraltar': 'EU', 'Greenland': 'NA', 'The Gambia': 'AF',
#                              'Guinea': 'AF', 'Guadeloupe': 'NA', 'Equatorial Guinea': 'AF', 'Greece': 'EU',
#                              'South Georgia and South Sandwich Islands': 'AN', 'Guatemala': 'NA', 'Guam': 'OC',
#                              'Guinea-Bissau': 'AF', 'Guyana': 'SA', 'Hong Kong': 'AS',
#                              'Heard and McDonald Islands': 'AN', 'Honduras': 'NA', 'Croatia': 'EU', 'Haiti': 'NA',
#                              'Hungary': 'EU', 'Indonesia': 'AS', 'Ireland': 'EU', 'Israel': 'AS', 'Isle of Man': 'EU',
#                              'India': 'AS', 'British Indian Ocean Territory': 'AS', 'Iraq': 'AS', 'Iran': 'AS',
#                              'Iceland': 'EU', 'Italy': 'EU', 'Jersey': 'EU', 'Jamaica': 'NA', 'Jordan': 'AS',
#                              'Japan': 'AS', 'Kenya': 'AF', 'Kyrgyzstan': 'AS', 'Cambodia': 'AS', 'Kiribati': 'OC',
#                              'Comoros': 'AF', 'St Kitts and Nevis': 'NA', 'North Korea': 'AS', 'South Korea': 'AS',
#                              'Kuwait': 'AS', 'Cayman Islands': 'NA', 'Kazakhstan': 'AS', 'Laos': 'AS', 'Lebanon': 'AS',
#                              'Saint Lucia': 'NA', 'Liechtenstein': 'EU', 'Sri Lanka': 'AS', 'Liberia': 'AF',
#                              'Lesotho': 'AF', 'Lithuania': 'EU', 'Luxembourg': 'EU', 'Latvia': 'EU', 'Libya': 'AF',
#                              'Morocco': 'AF', 'Monaco': 'EU', 'Moldova': 'EU', 'Montenegro': 'EU', 'Saint Martin': 'NA',
#                              'Madagascar': 'AF', 'Marshall Islands': 'OC', 'North Macedonia': 'EU', 'Mali': 'AF',
#                              'Myanmar': 'AS', 'Mongolia': 'AS', 'Macao': 'AS', 'Northern Mariana Islands': 'OC',
#                              'Martinique': 'NA', 'Mauritania': 'AF', 'Montserrat': 'NA', 'Malta': 'EU',
#                              'Mauritius': 'AF', 'Maldives': 'AS', 'Malawi': 'AF', 'Mexico': 'NA', 'Malaysia': 'AS',
#                              'Mozambique': 'AF', 'Namibia': 'AF', 'New Caledonia': 'OC', 'Niger': 'AF',
#                              'Norfolk Island': 'OC', 'Nigeria': 'AF', 'Nicaragua': 'NA', 'Netherlands': 'EU',
#                              'Norway': 'EU', 'Nepal': 'AS', 'Nauru': 'OC', 'Niue': 'OC', 'New Zealand': 'OC',
#                              'Oman': 'AS', 'Panama': 'NA', 'Peru': 'SA', 'French Polynesia': 'OC',
#                              'Papua New Guinea': 'OC', 'Philippines': 'AS', 'Pakistan': 'AS', 'Poland': 'EU',
#                              'Saint Pierre and Miquelon': 'NA', 'Pitcairn Islands': 'OC', 'Puerto Rico': 'NA',
#                              'Palestine': 'AS', 'Portugal': 'EU', 'Palau': 'OC', 'Paraguay': 'SA', 'Qatar': 'AS',
#                              'Réunion': 'AF', 'Romania': 'EU', 'Serbia': 'EU', 'Russia': 'EU', 'Rwanda': 'AF',
#                              'Saudi Arabia': 'AS', 'Solomon Islands': 'OC', 'Seychelles': 'AF', 'Sudan': 'AF',
#                              'Sweden': 'EU', 'Singapore': 'AS', 'Saint Helena': 'AF', 'Slovenia': 'EU',
#                              'Svalbard and Jan Mayen': 'EU', 'Slovakia': 'EU', 'Sierra Leone': 'AF', 'San Marino': 'EU',
#                              'Senegal': 'AF', 'Somalia': 'AF', 'Suriname': 'SA', 'South Sudan': 'AF',
#                              'São Tomé and Príncipe': 'AF', 'El Salvador': 'NA', 'Sint Maarten': 'NA', 'Syria': 'AS',
#                              'Eswatini': 'AF', 'Turks and Caicos Islands': 'NA', 'Chad': 'AF',
#                              'French Southern Territories': 'AN', 'Togo': 'AF', 'Thailand': 'AS', 'Tajikistan': 'AS',
#                              'Tokelau': 'OC', 'Timor-Leste': 'OC', 'Turkmenistan': 'AS', 'Tunisia': 'AF', 'Tonga': 'OC',
#                              'Turkey': 'AS', 'Trinidad and Tobago': 'NA', 'Tuvalu': 'OC', 'Taiwan': 'AS',
#                              'Tanzania': 'AF', 'Ukraine': 'EU', 'Uganda': 'AF', 'U.S. Outlying Islands': 'OC',
#                              'United States': 'NA', 'Uruguay': 'SA', 'Uzbekistan': 'AS', 'Vatican City': 'EU',
#                              'St Vincent and Grenadines': 'NA', 'Venezuela': 'SA', 'British Virgin Islands': 'NA',
#                              'U.S. Virgin Islands': 'NA', 'Vietnam': 'AS', 'Vanuatu': 'OC', 'Wallis and Futuna': 'OC',
#                              'Samoa': 'OC', 'Kosovo': 'EU', 'Yemen': 'AS', 'Mayotte': 'AF', 'South Africa': 'AF',
#                              'Zambia': 'AF', 'Zimbabwe': 'AF'}
# # def website_to_excel():
# #     driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# #     driver=wd.Chrome(executable_path=driver_path)
# #     driver.get("https://tradingeconomics.com/matrix")
# #     wait=WebDriverWait(driver,10)
# #     target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
# #     if os.path.exists("inflation.xlsx"):
# #         workbook=xl.load_workbook("inflation.xlsx")
# #     else:
# #         workbook=xl.Workbook()
# #     sheetnames=workbook.sheetnames
# #     sheetname="inflation"
# #     if sheetname in sheetnames:
# #         sheet_to_remove=workbook[sheetname]
# #         workbook.remove(sheet_to_remove)
# #     sheet1=workbook.create_sheet(sheetname)
# #     for i in target_tr:
# #         target_td=i.find_elements(By.TAG_NAME,"td")
# #         if len(target_td)>=9:
# #             target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
# #             target_gdp_a=target_td[1].find_element(By.TAG_NAME,"a")
# #             print(target_country_a.text, target_gdp_a.text)
# #             if target_country_a.text in country_continent_dict:
# #                 continent_name=country_continent_dict[target_country_a.text]
# #                 sheet1.append([target_country_a.text,target_gdp_a.text,continent_name])
# #                 print(target_country_a.text,target_gdp_a.text,continent_name)
#
#
#
# #
# #
# #
# #
# #
# #
# #     workbook.save("inflation.xlsx")
# #     driver.quit()
# # website_to_excel()
#
# def excel_to_graph():
#     import matplotlib.pyplot as plt
#     import openpyxl as xl
#     workbook=xl.load_workbook("inflation.xlsx")
#     sheet1=workbook["inflation"]
#     continent_gdp_dict={}
#     for i in sheet1.iter_rows(values_only=True):
#         continent=i[2]
#         gdp=i[1]
#         if continent not in continent_gdp_dict:
#             continent_gdp_dict[continent]=int(gdp)
#         else:
#             continent_gdp_dict[continent]=continent_gdp_dict[continent]+int(gdp)
#     continent_keys_list=list(continent_gdp_dict.keys())
#     gdp_values_list=list(continent_gdp_dict.values())
#     print(continent_keys_list)
#     print(gdp_values_list)
#
#     plt.plot(continent_keys_list,gdp_values_list)
#     plt.xlabel("continents")
#     plt.ylabel("gdp")
#     plt.title("gdp of continents")
#     plt.show()
#
import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as xl
import os
import matplotlib.pyplot as plt

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://data.worldbank.org/indicator/MS.MIL.XPND.GD.ZS?locations=AF")
# wait=WebDriverWait(driver,10)
# target_class=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"div.item:not([class *= ' '])")))
# dict_country_gdp_percent={}
# for i in target_class:
#     target_div=i.find_elements(By.TAG_NAME,"div")
#     target_country=target_div[0].text
#     target_gdp_percent=target_div[2].text
#     if target_gdp_percent==None or target_gdp_percent=="":
#         continue
#     else:
#         dict_country_gdp_percent[target_country] = target_gdp_percent
# print(dict_country_gdp_percent)
# dict_country_gdp_percent={'Afghanistan': '1.4', 'Albania': '1.4', 'Algeria': '5.6', 'Angola': '1.1', 'Argentina': '0.6', 'Armenia': '4.4', 'Australia': '2.0', 'Austria': '0.8', 'Azerbaijan': '5.3', 'Bahrain': '3.6', 'Bangladesh': '1.3', 'Belarus': '1.2', 'Belgium': '1.1', 'Belize': '1.3', 'Benin': '0.6', 'Bolivia': '1.5', 'Bosnia and Herzegovina': '0.9', 'Botswana': '2.9', 'Brazil': '1.2', 'Brunei Darussalam': '3.3', 'Bulgaria': '1.6', 'Burkina Faso': '2.9', 'Burundi': '1.9', 'Cabo Verde': '0.5', 'Cambodia': '2.3', 'Cameroon': '1.0', 'Canada': '1.3', 'Central African Republic': '1.9', 'Chad': '2.8', 'Chile': '2.0', 'China': '1.7', 'Colombia': '3.4', 'Congo, Dem. Rep.': '0.6', 'Congo, Rep.': '2.9', "Cote d'Ivoire": '1.3', 'Croatia': '2.7', 'Cuba': '2.9', 'Cyprus': '2.0', 'Czechia': '1.4', 'Denmark': '1.4', 'Djibouti': '3.7', 'Dominican Republic': '0.7', 'Ecuador': '2.5', 'Egypt, Arab Rep.': '1.3', 'El Salvador': '1.4', 'Equatorial Guinea': '1.3', 'Eritrea': '20.9', 'Estonia': '2.2', 'Eswatini': '1.7', 'Ethiopia': '0.5', 'Fiji': '1.7', 'Finland': '2.0', 'France': '1.9', 'Gabon': '1.7', 'Gambia, The': '0.7', 'Georgia': '1.7', 'Germany': '1.3', 'Ghana': '0.5', 'Greece': '3.9', 'Guatemala': '0.4', 'Guinea': '1.5', 'Guinea-Bissau': '1.8', 'Guyana': '1.0', 'Haiti': '0.2', 'Honduras': '1.6', 'Hungary': '1.6', 'India': '2.7', 'Indonesia': '0.7', 'Iran, Islamic Rep.': '2.3', 'Iraq': '2.3', 'Ireland': '0.3', 'Israel': '5.2', 'Italy': '1.5', 'Jamaica': '1.4', 'Japan': '1.1', 'Jordan': '5.0', 'Kazakhstan': '0.9', 'Kenya': '1.1', 'Korea, Rep.': '2.8', 'Kosovo': '0.9', 'Kuwait': '6.7', 'Kyrgyz Republic': '1.7', 'Lao PDR': '0.2', 'Latvia': '2.3', 'Lebanon': '3.0', 'Lesotho': '1.3', 'Liberia': '1.0', 'Libya': '15.5', 'Lithuania': '2.0', 'Luxembourg': '0.6', 'Madagascar': '0.8', 'Malawi': '0.9', 'Malaysia': '1.1', 'Mali': '2.8', 'Malta': '0.5', 'Mauritania': '2.4', 'Mauritius': '0.2', 'Mexico': '0.7', 'Moldova': '0.4', 'Mongolia': '0.7', 'Montenegro': '1.7', 'Morocco': '4.2', 'Mozambique': '1.5', 'Myanmar': '3.3', 'Namibia': '3.0', 'Nepal': '1.3', 'Netherlands': '1.4', 'New Zealand': '1.4', 'Nicaragua': '0.6', 'Niger': '1.8', 'Nigeria': '1.0', 'North Macedonia': '1.5', 'Norway': '1.8', 'Oman': '7.3', 'Pakistan': '3.8', 'Panama': '1.0', 'Papua New Guinea': '0.4', 'Paraguay': '1.0', 'Peru': '1.1', 'Philippines': '1.0', 'Poland': '2.1', 'Portugal': '2.1', 'Qatar': '4.8', 'Romania': '2.0', 'Russian Federation': '4.1', 'Rwanda': '1.5', 'Saudi Arabia': '6.6', 'Senegal': '1.8', 'Serbia': '2.1', 'Seychelles': '1.5', 'Sierra Leone': '0.6', 'Singapore': '3.0', 'Slovak Republic': '1.7', 'Slovenia': '1.2', 'Somalia': '1.5', 'South Africa': '0.9', 'South Sudan': '2.4', 'Spain': '1.4', 'Sri Lanka': '1.9', 'Sudan': '1.0', 'Sweden': '1.3', 'Switzerland': '0.7', 'Syrian Arab Republic': '4.1', 'Tajikistan': '1.0', 'Tanzania': '1.1', 'Thailand': '1.3', 'Timor-Leste': '1.1', 'Togo': '1.8', 'Trinidad and Tobago': '0.9', 'Tunisia': '2.8', 'Turkiye': '2.1', 'Turkmenistan': '2.9', 'Uganda': '2.4', 'Ukraine': '3.2', 'United Arab Emirates': '5.6', 'United Kingdom': '2.2', 'United States': '3.5', 'Uruguay': '2.3', 'Uzbekistan': '3.6', 'Venezuela, RB': '0.2', 'Vietnam': '2.3', 'Yemen, Rep.': '4.0', 'Zambia': '1.3', 'Zimbabwe': '0.0', 'World': '2.2', 'Arab World': '4.6', 'Caribbean small states': '1.0', 'Central Europe and the Baltics': '1.9', 'East Asia & Pacific': '1.7', 'East Asia & Pacific (excluding high income)': '1.6', 'Euro area': '1.5', 'Europe & Central Asia': '1.8', 'Europe & Central Asia (excluding high income)': '3.1', 'European Union': '1.5', 'Fragile and conflict affected situations': '1.6', 'Heavily indebted poor countries (HIPC)': '1.2', 'Latin America & Caribbean': '1.2', 'Latin America & Caribbean (excluding high income)': '1.1', 'Least developed countries: UN classification': '1.4', 'Middle East & North Africa': '4.4', 'Middle East & North Africa (excluding high income)': '2.7', 'North America': '3.3', 'OECD members': '2.4', 'Other small states': '3.3', 'Pacific island small states': '1.0', 'Small states': '3.1', 'South Asia': '2.6', 'Sub-Saharan Africa': '1.1', 'Sub-Saharan Africa (excluding high income)': '1.1', 'High income': '2.5', 'Low & middle income': '1.8', 'Low income': '1.2', 'Lower middle income': '2.1', 'Middle income': '1.8', 'Upper middle income': '1.8'}

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://tradingeconomics.com/matrix")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
# if os.path.exists("weapons_gdp.xlsx"):
#     workbook=xl.load_workbook("weapons_gdp.xlsx")
# else:
#     workbook=xl.Workbook()
# sheetnames=workbook.sheetnames
# sheetname="weapon_spending"
# if sheetname in sheetnames:
#     sheet_to_remove=workbook[sheetname]
#     workbook.remove((sheet_to_remove))
# sheet1=workbook.create_sheet(sheetname)
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
#     target_gdp_a = target_td[1].find_element(By.TAG_NAME, "a")
#     if target_country_a.text in dict_country_gdp_percent:
#         gdp_percent=float(dict_country_gdp_percent[target_country_a.text])
#         total_gdp_spending_weapons=int(target_gdp_a.text)*gdp_percent
#         sheet1.append([target_country_a.text,total_gdp_spending_weapons])
#         print(target_country_a.text, total_gdp_spending_weapons)
#     else:
#         continue
#
#
# workbook.save("weapons_gdp.xlsx")
# driver.quit()

# workbook=xl.load_workbook("weapons_gdp.xlsx")
# sheet1=workbook["weapon_spending"]
# country=[]
# dict_countries_gdp_spending={}
# print("choose 5 countries for which you want to see data")
# for i in range(5):
#     print("please choose your country")
#     choice_country=input("choose")
#     country.append(choice_country)
# for i in sheet1.iter_rows(values_only=True):
#     countries=i[0]
#     gdp_spending_weapons=i[1]
#     if countries in country:
#         dict_countries_gdp_spending[countries]=gdp_spending_weapons
#     else:
#         continue
# plt.bar(dict_countries_gdp_spending.keys(),dict_countries_gdp_spending.values())
# plt.xlabel("countries")
# plt.ylabel("GDP")
# plt.title("GDP spending on weapons")
# plt.show()

# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# import openpyxl as xl
# import os
# list=[]
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.geonames.org/countries/")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#countries tbody tr")))
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     target_country_a=target_td[4].find_element(By.TAG_NAME,"a")
#     target_continent=target_td[8].text
#     list.append([target_country_a.text,target_continent])
# print(type(target_country_a.text),type(target_continent))
# print(list)
# driver.quit()

import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as xl
import os
import matplotlib.pyplot as plt
# if os.path.exists("inflation.xlsx"):
#     workbook=xl.load_workbook("inflation.xlsx")
# else:
#     workbook=xl.Workbook()
# sheetnames=workbook.sheetnames
# sheetname="country_gdp"
# if sheetname in sheetnames:
#     sheet_to_remove=workbook[sheetname]
#     workbook.remove(sheet_to_remove)
# sheet1=workbook.create_sheet(sheetname)
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://tradingeconomics.com/matrix")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
#
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
#     target_gdp_a=target_td[1].find_element(By.TAG_NAME,"a")
#     sheet1.append([target_country_a.text,target_gdp_a.text])
# workbook.save("inflation.xlsx")
# driver.quit

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.geonames.org/countries/")
# wait=WebDriverWait(driver,10)
# dict_country_continent=[]
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#countries tbody tr")))
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     target_country_a=target_td[4].find_element(By.TAG_NAME,"a")
#     target_continent=target_td[8].text
#     dict_country_continent.append([target_country_a.text,target_continent])
# print(dict_country_continent)

dict_countries_continents = {'Andorra': 'EU', 'United Arab Emirates': 'AS', 'Afghanistan': 'AS',
                             'Antigua and Barbuda': 'NA', 'Anguilla': 'NA', 'Albania': 'EU', 'Armenia': 'AS',
                             'Angola': 'AF', 'Antarctica': 'AN', 'Argentina': 'SA', 'American Samoa': 'OC',
                             'Austria': 'EU', 'Australia': 'OC', 'Aruba': 'NA', 'Åland': 'EU', 'Azerbaijan': 'AS',
                             'Bosnia and Herzegovina': 'EU', 'Barbados': 'NA', 'Bangladesh': 'AS', 'Belgium': 'EU',
                             'Burkina Faso': 'AF', 'Bulgaria': 'EU', 'Bahrain': 'AS', 'Burundi': 'AF', 'Benin': 'AF',
                             'Saint Barthélemy': 'NA', 'Bermuda': 'NA', 'Brunei': 'AS', 'Bolivia': 'SA',
                             'Bonaire, Sint Eustatius, and Saba': 'NA', 'Brazil': 'SA', 'Bahamas': 'NA', 'Bhutan': 'AS',
                             'Bouvet Island': 'AN', 'Botswana': 'AF', 'Belarus': 'EU', 'Belize': 'NA', 'Canada': 'NA',
                             'Cocos (Keeling) Islands': 'AS', 'DR Congo': 'AF', 'Central African Republic': 'AF',
                             'Congo Republic': 'AF', 'Switzerland': 'EU', 'Ivory Coast': 'AF', 'Cook Islands': 'OC',
                             'Chile': 'SA', 'Cameroon': 'AF', 'China': 'AS', 'Colombia': 'SA', 'Costa Rica': 'NA',
                             'Cuba': 'NA', 'Cabo Verde': 'AF', 'Curaçao': 'NA', 'Christmas Island': 'OC',
                             'Cyprus': 'EU', 'Czechia': 'EU', 'Germany': 'EU', 'Djibouti': 'AF', 'Denmark': 'EU',
                             'Dominica': 'NA', 'Dominican Republic': 'NA', 'Algeria': 'AF', 'Ecuador': 'SA',
                             'Estonia': 'EU', 'Egypt': 'AF', 'Western Sahara': 'AF', 'Eritrea': 'AF', 'Spain': 'EU',
                             'Ethiopia': 'AF', 'Finland': 'EU', 'Fiji': 'OC', 'Falkland Islands': 'SA',
                             'Micronesia': 'OC', 'Faroe Islands': 'EU', 'France': 'EU', 'Gabon': 'AF',
                             'United Kingdom': 'EU', 'Grenada': 'NA', 'Georgia': 'AS', 'French Guiana': 'SA',
                             'Guernsey': 'EU', 'Ghana': 'AF', 'Gibraltar': 'EU', 'Greenland': 'NA', 'The Gambia': 'AF',
                             'Guinea': 'AF', 'Guadeloupe': 'NA', 'Equatorial Guinea': 'AF', 'Greece': 'EU',
                             'South Georgia and South Sandwich Islands': 'AN', 'Guatemala': 'NA', 'Guam': 'OC',
                             'Guinea-Bissau': 'AF', 'Guyana': 'SA', 'Hong Kong': 'AS',
                             'Heard and McDonald Islands': 'AN', 'Honduras': 'NA', 'Croatia': 'EU', 'Haiti': 'NA',
                             'Hungary': 'EU', 'Indonesia': 'AS', 'Ireland': 'EU', 'Israel': 'AS', 'Isle of Man': 'EU',
                             'India': 'AS', 'British Indian Ocean Territory': 'AS', 'Iraq': 'AS', 'Iran': 'AS',
                             'Iceland': 'EU', 'Italy': 'EU', 'Jersey': 'EU', 'Jamaica': 'NA', 'Jordan': 'AS',
                             'Japan': 'AS', 'Kenya': 'AF', 'Kyrgyzstan': 'AS', 'Cambodia': 'AS', 'Kiribati': 'OC',
                             'Comoros': 'AF', 'St Kitts and Nevis': 'NA', 'North Korea': 'AS', 'South Korea': 'AS',
                             'Kuwait': 'AS', 'Cayman Islands': 'NA', 'Kazakhstan': 'AS', 'Laos': 'AS', 'Lebanon': 'AS',
                             'Saint Lucia': 'NA', 'Liechtenstein': 'EU', 'Sri Lanka': 'AS', 'Liberia': 'AF',
                             'Lesotho': 'AF', 'Lithuania': 'EU', 'Luxembourg': 'EU', 'Latvia': 'EU', 'Libya': 'AF',
                             'Morocco': 'AF', 'Monaco': 'EU', 'Moldova': 'EU', 'Montenegro': 'EU', 'Saint Martin': 'NA',
                             'Madagascar': 'AF', 'Marshall Islands': 'OC', 'North Macedonia': 'EU', 'Mali': 'AF',
                             'Myanmar': 'AS', 'Mongolia': 'AS', 'Macao': 'AS', 'Northern Mariana Islands': 'OC',
                             'Martinique': 'NA', 'Mauritania': 'AF', 'Montserrat': 'NA', 'Malta': 'EU',
                             'Mauritius': 'AF', 'Maldives': 'AS', 'Malawi': 'AF', 'Mexico': 'NA', 'Malaysia': 'AS',
                             'Mozambique': 'AF', 'Namibia': 'AF', 'New Caledonia': 'OC', 'Niger': 'AF',
                             'Norfolk Island': 'OC', 'Nigeria': 'AF', 'Nicaragua': 'NA', 'Netherlands': 'EU',
                             'Norway': 'EU', 'Nepal': 'AS', 'Nauru': 'OC', 'Niue': 'OC', 'New Zealand': 'OC',
                             'Oman': 'AS', 'Panama': 'NA', 'Peru': 'SA', 'French Polynesia': 'OC',
                             'Papua New Guinea': 'OC', 'Philippines': 'AS', 'Pakistan': 'AS', 'Poland': 'EU',
                             'Saint Pierre and Miquelon': 'NA', 'Pitcairn Islands': 'OC', 'Puerto Rico': 'NA',
                             'Palestine': 'AS', 'Portugal': 'EU', 'Palau': 'OC', 'Paraguay': 'SA', 'Qatar': 'AS',
                             'Réunion': 'AF', 'Romania': 'EU', 'Serbia': 'EU', 'Russia': 'EU', 'Rwanda': 'AF',
                             'Saudi Arabia': 'AS', 'Solomon Islands': 'OC', 'Seychelles': 'AF', 'Sudan': 'AF',
                             'Sweden': 'EU', 'Singapore': 'AS', 'Saint Helena': 'AF', 'Slovenia': 'EU',
                             'Svalbard and Jan Mayen': 'EU', 'Slovakia': 'EU', 'Sierra Leone': 'AF', 'San Marino': 'EU',
                             'Senegal': 'AF', 'Somalia': 'AF', 'Suriname': 'SA', 'South Sudan': 'AF',
                             'São Tomé and Príncipe': 'AF', 'El Salvador': 'NA', 'Sint Maarten': 'NA', 'Syria': 'AS',
                             'Eswatini': 'AF', 'Turks and Caicos Islands': 'NA', 'Chad': 'AF',
                             'French Southern Territories': 'AN', 'Togo': 'AF', 'Thailand': 'AS', 'Tajikistan': 'AS',
                             'Tokelau': 'OC', 'Timor-Leste': 'OC', 'Turkmenistan': 'AS', 'Tunisia': 'AF', 'Tonga': 'OC',
                             'Turkey': 'AS', 'Trinidad and Tobago': 'NA', 'Tuvalu': 'OC', 'Taiwan': 'AS',
                             'Tanzania': 'AF', 'Ukraine': 'EU', 'Uganda': 'AF', 'U.S. Outlying Islands': 'OC',
                             'United States': 'NA', 'Uruguay': 'SA', 'Uzbekistan': 'AS', 'Vatican City': 'EU',
                             'St Vincent and Grenadines': 'NA', 'Venezuela': 'SA', 'British Virgin Islands': 'NA',
                             'U.S. Virgin Islands': 'NA', 'Vietnam': 'AS', 'Vanuatu': 'OC', 'Wallis and Futuna': 'OC',
                             'Samoa': 'OC', 'Kosovo': 'EU', 'Yemen': 'AS', 'Mayotte': 'AF', 'South Africa': 'AF',
                             'Zambia': 'AF', 'Zimbabwe': 'AF'}

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://tradingeconomics.com/matrix")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
# if os.path.exists("inflation.xlsx"):
#     workbook=xl.load_workbook("inflation.xlsx")
# else:
#     workbook=xl.Workbook()
# sheetnames=workbook.sheetnames
# sheetname="country_gdp"
# if sheetname in sheetnames:
#     sheet_to_remove=workbook[sheetname]
#     workbook.remove(sheet_to_remove)
# sheet1=workbook.create_sheet(sheetname)
# for i in target_tr:
#     target_td = i.find_elements(By.TAG_NAME, "td")
#     if len(target_td)>=9:
#         target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
#         target_gdp_a=target_td[1].find_element(By.TAG_NAME,"a")
#         if target_country_a.text in dict_countries_continents:
#             continents=dict_countries_continents[target_country_a.text]
#             sheet1.append([target_country_a.text,target_gdp_a.text,continents])
# workbook.save("inflation.xlsx")
# driver.quit()

# workbook=xl.load_workbook("inflation.xlsx")
# sheet1=workbook["country_gdp"]
# continent_dict={}
#
#
# for i in sheet1.iter_rows(values_only=True):
#     continents=i[2]
#     gdp=i[1]
#     if continents not in continent_dict:
#         continent_dict[continents]=int(gdp)
#     else:
#         continent_dict[continents]=continent_dict[continents]+int(gdp)
# plt.pie(continent_dict.values(),labels=continent_dict.keys(),autopct="%1.1f%%")
# plt.xlabel("continents")
# plt.ylabel("gdp")
# plt.show()
#
# import selenium.webdriver as wd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# import openpyxl
# import os
# import matplotlib.pyplot as plt
#
# # driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# # driver=wd.Chrome(executable_path=driver_path)
# # driver.get("https://www.geonames.org/countries/")
# # wait=WebDriverWait(driver,10)
# # target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#countries tbody tr")))
# # if os.path.exists("inflation.xlsx"):
# #     workbook=xl.load_workbook("inflation.xlsx")
# # else:
# #     workbook=xl.Workbook()
# # sheentames=workbook.sheetnames
# # sheetname="continent_gdp"
# # if sheetname in sheentames:
# #     sheet_to_remove=workbook[sheetname]
# #     workbook.remove(sheet_to_remove)
# # sheet1=workbook.create_sheet(sheetname)
# # for i in target_tr:
# #     target_td=i.find_elements(By.TAG_NAME,"td")
# #     if len(target_td)>=9:
# #         target_country_a=target_td[4].find_element(By.TAG_NAME,"a")
# #         target_continent=target_td[8].text
# #         sheet1.append([target_country_a.text,target_continent])
# # workbook.save("inflation.xlsx")
# # driver.quit()
#
# # driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# # driver=wd.Chrome(executable_path=driver_path)
# # driver.get("https://tradingeconomics.com/matrix")
# # wait=WebDriverWait(driver,10)
# # dict_country_gdp={}
# # target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
# # for i in target_tr:
# #     target_td=i.find_elements(By.TAG_NAME,"td")
# #     target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
# #     target_gdp_a=target_td[1].find_element(By.TAG_NAME,"a")
# #     dict_country_gdp[target_country_a.text]=target_gdp_a.text
# # print(dict_country_gdp)
# # dict_country_gdp={'United States': '23315', 'China': '17734', 'Euro Area': '14563', 'Japan': '4941', 'Germany': '4260', 'India': '3176', 'United Kingdom': '3131', 'France': '2958', 'Italy': '2108', 'Canada': '1988', 'South Korea': '1799', 'Russia': '1776', 'Brazil': '1609', 'Australia': '1553', 'Spain': '1427', 'Mexico': '1273', 'Indonesia': '1186', 'Netherlands': '1013', 'Saudi Arabia': '834', 'Turkey': '819', 'Switzerland': '801', 'Taiwan': '775', 'Poland': '679', 'Sweden': '636', 'Belgium': '594', 'Thailand': '506', 'Ireland': '504', 'Israel': '489', 'Argentina': '487', 'Norway': '482', 'Austria': '480', 'Nigeria': '441', 'South Africa': '419', 'Bangladesh': '416', 'United Arab Emirates': '415', 'Egypt': '404', 'Denmark': '398', 'Singapore': '397', 'Philippines': '394', 'Malaysia': '373', 'Hong Kong': '368', 'Vietnam': '366', 'Pakistan': '348', 'Chile': '317', 'Colombia': '314', 'Finland': '297', 'Romania': '284', 'Czech Republic': '282', 'Portugal': '254', 'New Zealand': '250', 'Iran': '232', 'Peru': '223', 'Greece': '215', 'Iraq': '208', 'Ukraine': '200', 'Kazakhstan': '197', 'Hungary': '182', 'Qatar': '180', 'Algeria': '163', 'Morocco': '143', 'Slovakia': '115', 'Venezuela': '112', 'Ethiopia': '111', 'Kenya': '110', 'Puerto Rico': '107', 'Cuba': '107', 'Ecuador': '106', 'Kuwait': '106', 'Dominican Republic': '94', 'Sri Lanka': '89', 'Oman': '88', 'Luxembourg': '86', 'Guatemala': '86', 'Bulgaria': '84', 'Ghana': '78', 'Ivory Coast': '70', 'Croatia': '69', 'Uzbekistan': '69', 'Tanzania': '68', 'Belarus': '68', 'Angola': '67', 'Lithuania': '66', 'Myanmar': '65', 'Syria': '65', 'Panama': '64', 'Costa Rica': '64', 'Serbia': '63', 'Slovenia': '62', 'Uruguay': '59', 'Azerbaijan': '55', 'Congo': '54', 'Tunisia': '47', 'Jordan': '46', 'Cameroon': '45', 'Turkmenistan': '45', 'Libya': '43', 'Uganda': '41', 'Latvia': '40', 'Paraguay': '40', 'Bolivia': '40', 'Bahrain': '39', 'Estonia': '37', 'Nepal': '36', 'Sudan': '34', 'El Salvador': '32', 'Macau': '30', 'Honduras': '28', 'Senegal': '28', 'Cyprus': '28', 'Zimbabwe': '28', 'Cambodia': '27', 'Papua New Guinea': '27', 'Iceland': '26', 'Trinidad and Tobago': '24', 'Bosnia and Herzegovina': '23', 'Lebanon': '23', 'Zambia': '22', 'Yemen': '21', 'Haiti': '21', 'Gabon': '20', 'Burkina Faso': '20', 'Georgia': '19', 'Mali': '19', 'Laos': '19', 'Palestine': '18', 'Botswana': '18', 'Albania': '18', 'Benin': '17', 'Malta': '17', 'Mozambique': '16', 'Guinea': '16', 'Afghanistan': '15', 'Mongolia': '15', 'Niger': '15', 'Jamaica': '15', 'Moldova': '14', 'Macedonia': '14', 'Madagascar': '14', 'Nicaragua': '14', 'Armenia': '14', 'Brunei': '14', 'Republic of the Congo': '13', 'Malawi': '13', 'Mauritius': '12', 'Namibia': '12', 'Chad': '12', 'Equatorial Guinea': '12', 'Bahamas': '11', 'Rwanda': '11', 'Mauritania': '10', 'Monaco': '9', 'Kyrgyzstan': '9', 'Kosovo': '9', 'New Caledonia': '9', 'Tajikistan': '9', 'Togo': '8', 'Somalia': '8', 'Guyana': '8', 'Cayman Islands': '6', 'Montenegro': '6', 'Liechtenstein': '6', 'Maldives': '5', 'Swaziland': '5', 'Sierra Leone': '4', 'Liberia': '4', 'Fiji': '4', 'Djibouti': '3', 'Central African Republic': '3', 'Burundi': '3', 'Bhutan': '3', 'Lesotho': '3', 'Suriname': '3', 'Cape Verde': '2', 'East Timor': '2', 'Eritrea': '2', 'Gambia': '2', 'Guinea Bissau': '2', 'Comoros': '1', 'Sao Tome and Principe': '1', 'South Sudan': '1', 'Seychelles': '1'}
# #
# # driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# # driver=wd.Chrome(executable_path=driver_path)
# # driver.get("https://www.geonames.org/countries/")
# #
# # wait=WebDriverWait(driver,10)
# # target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#countries tbody tr")))
# # if os.path.exists("inflation.xlsx"):
# #     workbook=xl.load_workbook("inflation.xlsx")
# # else:
# #     workbook=xl.Workbook()
# # sheetnames=workbook.sheetnames
# # sheetname="continent_gdp"
# # if sheetname in sheetnames:
# #     sheet_to_remove=workbook[sheetname]
# #     workbook.remove(sheet_to_remove)
# # sheet1=workbook.create_sheet(sheetname)
# # for i in target_tr:
# #     target_td=i.find_elements(By.TAG_NAME,"td")
# #     if len(target_td)>=9:
# #         target_country_a=target_td[4].find_element(By.TAG_NAME,"a")
# #         target_continent_a=target_td[8].text
# #         if target_country_a.text in dict_country_gdp:
# #             gdp_value=dict_country_gdp[target_country_a.text]
# #             sheet1.append([target_country_a.text,target_continent_a,gdp_value])
# #         else:
# #             continue
# #
# # workbook.save("inflation.xlsx")
# # driver.quit
# workbook=xl.load_workbook("inflation.xlsx")
# sheet1=workbook["continent_gdp"]
# dict_continents_gdp={}
# for i in sheet1.iter_rows(values_only=True):
#     countries=i[0]
#     continents=i[1]
#     gdp=i[2]
#
#     if continents not in dict_continents_gdp:
#         dict_continents_gdp[continents]=int(gdp)
#     else:
#         dict_continents_gdp[continents]=dict_continents_gdp[continents]+int(gdp)
# print(dict_continents_gdp)
# plt.bar(dict_continents_gdp.keys(),dict_continents_gdp.values())
# plt.xlabel("continents")
# plt.ylabel("gdp")
# plt.title("gdp of continents")
# plt.show()

import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as xl
import os
import matplotlib.pyplot as plt

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.numbeo.com/cost-of-living/rankings.jsp")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#t2 tbody tr")))
# if os.path.exists("cost_of_living.xlsx"):
#     workbook=xl.load_workbook("inflation.xlsx")
# else:
#     workbook=xl.Workbook()
# sheetnames=workbook.sheetnames
# sheetname="cost_of_living"
# if sheetname in sheetnames:
#     sheet_to_remove=workbook[sheetname]
#     workbook.remove(sheet_to_remove)
# sheet1=workbook.create_sheet(sheetname)
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     target_city_country=target_td[1].text
#     new_target_city_country=target_city_country.split(",")
#     print(new_target_city_country)
#     selected_city=new_target_city_country[0]
#     selected_country=new_target_city_country[len(new_target_city_country)-1]
#     target_cost_of_living_index=target_td[2].text
#     # print(selected_city,selected_country,target_cost_of_living_index)
#     sheet1.append([selected_city,selected_country,target_cost_of_living_index])
# workbook.save("cost_of_living.xlsx")
# driver.quit()

# workbook=xl.load_workbook("cost_of_living.xlsx")
# sheet1=workbook["cost_of_living"]
# dict_city_cost_living={}
# print("please choose your country")
# chosen_country=input("please choose your desired country")
# print(chosen_country)
# for i in sheet1.iter_rows(values_only=True):
#     cities=i[0]
#     countries=i[1].strip()
#     cost_of_living_index=float(i[2])
#     print(cities,countries,cost_of_living_index)
#     if countries==chosen_country:
#         dict_city_cost_living[cities]=cost_of_living_index
#     else:
#         continue
# plt.plot(dict_city_cost_living.keys(),dict_city_cost_living.values(),marker="*",ms=5,ls="dotted")
# plt.xlabel("cities")
# plt.ylabel("cost_of_living")
# plt.title("cost_of_living_of_cities")
# plt.show()
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://www.numbeo.com/property-investment/rankings.jsp")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.TAG_NAME,"tr")))
# if os.path.exists("cost_of_living.xlsx"):
#     workbook=xl.load_workbook("cost_of_living.xlsx")
# else:
#     workbook=xl.Workbook()
# sheetnames=workbook.sheetnames
# sheetname="price_to_income_ratio"
# if sheetname in sheetnames:
#     sheet_to_remove=workbook[sheetname]
#     workbook.remove(sheet_to_remove)
# sheet1=workbook.create_sheet(sheetname)
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     if len(target_td)>=9:
#         target_city_country=target_td[1].text.split(",")
#         specified_city=target_city_country[0]
#         specified_country=target_city_country[len(target_city_country)-1]
#         target_price_to_income_ratio=target_td[2].text
#         print(specified_city,specified_country,target_price_to_income_ratio)
#         sheet1.append([specified_city,specified_country,target_price_to_income_ratio])
# workbook.save("cost_of_living.xlsx")
# driver.quit()
#
# workbook=xl.load_workbook("cost_of_living.xlsx")
# sheet1=workbook["price_to_income_ratio"]
# dict_city_price_income_ratio={}
# print("please choose a country")
# country_chosen=input("please choose country")
# for i in sheet1.iter_rows(values_only=True):
#     cities=i[0]
#     country=i[1].strip()
#     price_to_income_rato=float(i[2])
#     if country==country_chosen:
#         dict_city_price_income_ratio[cities]=price_to_income_rato
#     else:
#         continue
# plt.bar(dict_city_price_income_ratio.keys(),dict_city_price_income_ratio.values())
# plt.xlabel("cities")
# plt.ylabel("price to income ratio")
# plt.title("cities price to income ratio")
# plt.show()

# import selenium.webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as xl
import os
import matplotlib.pyplot
#
# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://tradingeconomics.com/matrix")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
# if os.path.exists("weapons_gdp.xlsx"):
#     workbook=xl.load_workbook("weapons_gdp.xlsx")
# else:
#     workbook=xl.Workbook()
# sheetnames=workbook.sheetnames
# sheetname="weapons_spending"
# if sheetname in sheetnames:
#     sheetname_to_remove=workbook[sheetname]
#     workbook.remove(sheetname_to_remove)
# sheet1=workbook.create_sheet(sheetname)
# for i in target_tr:
#     target_td=i.find_elements(By.TAG_NAME,"td")
#     target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
#     target_gdp_a=target_td[1].find_element(By.TAG_NAME,"a")
#     print(target_country_a.text)
#     print(target_gdp_a.text)
#     sheet1.append([target_country_a.text,target_gdp_a.text])
# workbook.save("weapons_gdp.xlsx")
# driver.quit()

# driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
# driver=wd.Chrome(executable_path=driver_path)
# driver.get("https://data.worldbank.org/indicator/MS.MIL.XPND.GD.ZS?locations=AF")
# wait=WebDriverWait(driver,10)
# target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"div.item:not([class *= ' ']")))
# dict_country_gdp={}
# for i in target_tr:
#     target_div=i.find_elements(By.TAG_NAME,"div")
#     if len(target_div)>=4:
#         target_country=target_div[0].text
#         target_gdp_percent=target_div[2].text
#         print(target_country)
#         print(target_gdp_percent)
#     dict_country_gdp[target_country_div]=target_gdp_percent_div
# print(dict_country_gdp)

driver_path="C:/Users/Admin/Desktop/pythonpractice/chromedriver.exe"
driver=wd.Chrome(executable_path=driver_path)
driver.get("https://tradingeconomics.com/matrix")
wait=WebDriverWait(driver,10)
target_tr=wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,"table#matrix tbody tr")))
dict_country_gdp_percent={'Afghanistan': '1.4', 'Albania': '1.4', 'Algeria': '5.6', 'Angola': '1.1', 'Argentina': '0.6', 'Armenia': '4.4', 'Australia': '2.0', 'Austria': '0.8', 'Azerbaijan': '5.3', 'Bahrain': '3.6', 'Bangladesh': '1.3', 'Belarus': '1.2', 'Belgium': '1.1', 'Belize': '1.3', 'Benin': '0.6', 'Bolivia': '1.5', 'Bosnia and Herzegovina': '0.9', 'Botswana': '2.9', 'Brazil': '1.2', 'Brunei Darussalam': '3.3', 'Bulgaria': '1.6', 'Burkina Faso': '2.9', 'Burundi': '1.9', 'Cabo Verde': '0.5', 'Cambodia': '2.3', 'Cameroon': '1.0', 'Canada': '1.3', 'Central African Republic': '1.9', 'Chad': '2.8', 'Chile': '2.0', 'China': '1.7', 'Colombia': '3.4', 'Congo, Dem. Rep.': '0.6', 'Congo, Rep.': '2.9', "Cote d'Ivoire": '1.3', 'Croatia': '2.7', 'Cuba': '2.9', 'Cyprus': '2.0', 'Czechia': '1.4', 'Denmark': '1.4', 'Djibouti': '3.7', 'Dominican Republic': '0.7', 'Ecuador': '2.5', 'Egypt, Arab Rep.': '1.3', 'El Salvador': '1.4', 'Equatorial Guinea': '1.3', 'Eritrea': '20.9', 'Estonia': '2.2', 'Eswatini': '1.7', 'Ethiopia': '0.5', 'Fiji': '1.7', 'Finland': '2.0', 'France': '1.9', 'Gabon': '1.7', 'Gambia, The': '0.7', 'Georgia': '1.7', 'Germany': '1.3', 'Ghana': '0.5', 'Greece': '3.9', 'Guatemala': '0.4', 'Guinea': '1.5', 'Guinea-Bissau': '1.8', 'Guyana': '1.0', 'Haiti': '0.2', 'Honduras': '1.6', 'Hungary': '1.6', 'India': '2.7', 'Indonesia': '0.7', 'Iran, Islamic Rep.': '2.3', 'Iraq': '2.3', 'Ireland': '0.3', 'Israel': '5.2', 'Italy': '1.5', 'Jamaica': '1.4', 'Japan': '1.1', 'Jordan': '5.0', 'Kazakhstan': '0.9', 'Kenya': '1.1', 'Korea, Rep.': '2.8', 'Kosovo': '0.9', 'Kuwait': '6.7', 'Kyrgyz Republic': '1.7', 'Lao PDR': '0.2', 'Latvia': '2.3', 'Lebanon': '3.0', 'Lesotho': '1.3', 'Liberia': '1.0', 'Libya': '15.5', 'Lithuania': '2.0', 'Luxembourg': '0.6', 'Madagascar': '0.8', 'Malawi': '0.9', 'Malaysia': '1.1', 'Mali': '2.8', 'Malta': '0.5', 'Mauritania': '2.4', 'Mauritius': '0.2', 'Mexico': '0.7', 'Moldova': '0.4', 'Mongolia': '0.7', 'Montenegro': '1.7', 'Morocco': '4.2', 'Mozambique': '1.5', 'Myanmar': '3.3', 'Namibia': '3.0', 'Nepal': '1.3', 'Netherlands': '1.4', 'New Zealand': '1.4', 'Nicaragua': '0.6', 'Niger': '1.8', 'Nigeria': '1.0', 'North Macedonia': '1.5', 'Norway': '1.8', 'Oman': '7.3', 'Pakistan': '3.8', 'Panama': '1.0', 'Papua New Guinea': '0.4', 'Paraguay': '1.0', 'Peru': '1.1', 'Philippines': '1.0', 'Poland': '2.1', 'Portugal': '2.1', 'Qatar': '4.8', 'Romania': '2.0', 'Russian Federation': '4.1', 'Rwanda': '1.5', 'Saudi Arabia': '6.6', 'Senegal': '1.8', 'Serbia': '2.1', 'Seychelles': '1.5', 'Sierra Leone': '0.6', 'Singapore': '3.0', 'Slovak Republic': '1.7', 'Slovenia': '1.2', 'Somalia': '1.5', 'South Africa': '0.9', 'South Sudan': '2.4', 'Spain': '1.4', 'Sri Lanka': '1.9', 'Sudan': '1.0', 'Sweden': '1.3', 'Switzerland': '0.7', 'Syrian Arab Republic': '4.1', 'Tajikistan': '1.0', 'Tanzania': '1.1', 'Thailand': '1.3', 'Timor-Leste': '1.1', 'Togo': '1.8', 'Trinidad and Tobago': '0.9', 'Tunisia': '2.8', 'Turkiye': '2.1', 'Turkmenistan': '2.9', 'Uganda': '2.4', 'Ukraine': '3.2', 'United Arab Emirates': '5.6', 'United Kingdom': '2.2', 'United States': '3.5', 'Uruguay': '2.3', 'Uzbekistan': '3.6', 'Venezuela, RB': '0.2', 'Vietnam': '2.3', 'Yemen, Rep.': '4.0', 'Zambia': '1.3', 'Zimbabwe': '0.0', 'World': '2.2', 'Arab World': '4.6', 'Caribbean small states': '1.0', 'Central Europe and the Baltics': '1.9', 'East Asia & Pacific': '1.7', 'East Asia & Pacific (excluding high income)': '1.6', 'Euro area': '1.5', 'Europe & Central Asia': '1.8', 'Europe & Central Asia (excluding high income)': '3.1', 'European Union': '1.5', 'Fragile and conflict affected situations': '1.6', 'Heavily indebted poor countries (HIPC)': '1.2', 'Latin America & Caribbean': '1.2', 'Latin America & Caribbean (excluding high income)': '1.1', 'Least developed countries: UN classification': '1.4', 'Middle East & North Africa': '4.4', 'Middle East & North Africa (excluding high income)': '2.7', 'North America': '3.3', 'OECD members': '2.4', 'Other small states': '3.3', 'Pacific island small states': '1.0', 'Small states': '3.1', 'South Asia': '2.6', 'Sub-Saharan Africa': '1.1', 'Sub-Saharan Africa (excluding high income)': '1.1', 'High income': '2.5', 'Low & middle income': '1.8', 'Low income': '1.2', 'Lower middle income': '2.1', 'Middle income': '1.8', 'Upper middle income': '1.8'}
dict_haris={}
country_choice=[]
if os.path.exists("weapons_gdp.xlsx"):
    workbook=xl.load_workbook("weapons_gdp.xlsx")
else:
    workbook=xl.Workbook()
sheetnames=workbook.sheetnames
sheetname="weapons_spending"
if sheetname in sheetnames:
    sheet_to_remove=workbook[sheetname]
    workbook.remove(sheet_to_remove)
sheet1=workbook.create_sheet(sheetname)
for i in target_tr:
    target_td=i.find_elements(By.TAG_NAME,"td")
    target_country_a=target_td[0].find_element(By.TAG_NAME,"a")
    target_gdp_a=target_td[1].find_element(By.TAG_NAME,"a")
    if target_country_a.text in dict_country_gdp_percent:
        gdp_value=float(dict_country_gdp_percent[target_country_a.text])
        actual_gdp_value=gdp_value*(int(target_gdp_a.text)/100)
    else:
        continue
    sheet1.append([target_country_a.text,actual_gdp_value])

workbook.save("weapons_gdp.xlsx")
driver.quit()

# workbook=xl.load_workbook("weapons_gdp.xlsx")
# sheet1=workbook["weapons_spending"]
# choice_country_list=[]
# dict_country_gdp_weapons={}
# print("choose 5 countries for which you want to see data")
# for i in range(5):
#     choice_country=input("choose your country")
#     choice_country_list.append(choice_country)
# for i in sheet1.iter_rows(values_only=True):
#     country=i[0]
#     gdp_spending_weapons=i[1]
#     if country in choice_country_list:
#         dict_country_gdp_weapons[country]=gdp_spending_weapons
#     else:
#         continue
# print(dict_country_gdp_weapons)
# plt.bar(dict_country_gdp_weapons.keys(),dict_country_gdp_weapons.values())
# plt.xlabel("countries")
# plt.ylabel("GDP spending on weapons")
# plt.title("Total GDP spending on weapons by different nations")
# plt.show()









