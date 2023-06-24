import openpyxl as xl
import os






def library_management():
    if os.path.exists("student_python.xlsx"):
        workbook = xl.load_workbook("student_python.xlsx")
    else:
        workbook = xl.Workbook()
    sheet1=workbook.active
    books_add=int(input("how many books do you want to enter"))
    for i in range(books_add):
        name=input("name")
        author=input("author")
        sheet1.append([name,author])
    workbook.save("student_python.xlsx")
def event_management():
    if os.path.exists("student_python.xlsx"):
        workbook = xl.load_workbook("student_python.xlsx")
    else:
        workbook = xl.Workbook()
    sheet2=workbook.create_sheet("hello weddings")
    events_add=int(input("how many events do you want to enter"))
    for i in range(events_add):
        name=input("event name")
        guests=input("name guests")
        sheet2.append([name,guests])
    workbook.save("student_python.xlsx")

print("1 for sm")
print("2 for evm")
choice=int(input("dp your chouce"))

if choice==1:
    library_management()
elif choice==2:
    event_management()


def library_management():
    sheet1=workbook.active
    books_add=int(input("how many books do you want to add"))
    for i in range(books_add):
        name=input("what is the name of the book")
        author=input("what is name of author")
        sheet1.append([name,author])
    workbook.save("student_python.xlsx")

def student_management():
    sheet2=workbook.create_sheet("studentsd sheet")
    students_add=int(input("how many students do you want to add"))
    for i in range(students_add):
        name=input("what is name of student")
        classe=int(input("which class he studies in"))
        sheet2.append([name,classe])
    workbook.save("student_python.xlsx")

print("press 1 for library management")
print("press 2 for student management")

choice=int(input("input your choice"))

if choice==1:
    library_management()
elif choice==2:
    student_management()
