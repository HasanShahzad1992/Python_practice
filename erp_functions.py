import openpyxl as xl
import os


def library_management_add():
    if os.path.exists("pak_nz.xlsx"):
        workbook = xl.load_workbook("pak_nz.xlsx")
    else:
        workbook = xl.Workbook()
    sheet1 = workbook.active
    print(sheet1.max_row)
    add_books = int(input("please state the number of books that you want to add"))
    for i in range(add_books):
        book_id = sheet1.max_row
        name_book = input("please tell the name of the book")
        author_book = input("please tell the name of the author")
        year_publication = int(input("tell the name of the year in which the book was published"))
        sheet1.append([book_id,name_book, author_book, year_publication])
    workbook.save("pak_nz.xlsx")


def event_management_add():
    if os.path.exists("pak_nz.xlsx"):
        workbook = xl.load_workbook("pak_nz.xlsx")
    else:
        workbook = xl.Workbook()
    sheet_names = workbook.sheetnames
    sheet_name = "event_sheet"
    if sheet_name in sheet_names:
        sheet2 = workbook["event_sheet"]
    else:
        sheet2 = workbook.create_sheet("event_sheet")
    add_events = int(input("please tell how many events do you want to add"))
    for i in range(add_events):
        event_id=sheet2.max_row
        type_event = input("please tell the type of event you want to attend")
        guests = int(input("tell how many guests would enter into the function"))
        expenditure = int(input("what would be your expected expenditure"))
        sheet2.append([event_id,type_event, guests, expenditure])
    workbook.save("pak_nz.xlsx")


def student_management_add():
    if os.path.exists("pak_nz.xlsx"):
        workbook = xl.load_workbook("pak_nz.xlsx")
    else:
        workbook = xl.Workbook()
    sheet_names = workbook.sheetnames
    sheet_name = "student_sheet"
    if sheet_name in sheet_names:
        sheet3 = workbook["student_sheet"]
    else:
        sheet3 = workbook.create_sheet("student_sheet")
    add_students = int(input("tell how many students do you want to add"))
    for i in range(add_students):
        student_id=sheet3.max_row
        name_student = input("please tell the name of the student")
        grades = int(input("tell the marks he has achieved in HSSC"))
        college = input("please tell from which college has he passed his HSSC exams")
        sheet3.append([student_id,name_student, grades, college])
    workbook.save("pak_nz.xlsx")


def employee_management_add():
    if os.path.exists("pak_nz.xlsx"):
        workbook = xl.load_workbook("pak_nz.xlsx")
    else:
        workbook = xl.Workbook()
    sheet_names = workbook.sheetnames
    print(sheet_names)
    sheet_name = "employee_sheet"
    if sheet_name in sheet_names:
        sheet4 = workbook["employee_sheet"]
    else:
        sheet4 = workbook.create_sheet("employee_sheet")
    add_employee = int(input("please tell how many employees do you want to add"))
    for i in range(add_employee):
        employee_id=sheet4.max_row
        name_employee = input("what is the name of the employee")
        experience = float(input("tell the years of experience"))
        organization = input("what is the organization name")
        sheet4.append([employee_id,name_employee, experience, organization])
    workbook.save("pak_nz.xlsx")


def inventory_management_add():
    if os.path.exists("pak_nz.xlsx"):
        workbook = xl.load_workbook("pak_nz.xlsx")
    else:
        workbook = xl.Workbook()
    sheet_names = workbook.sheetnames
    sheet_name = "inventory_sheet"
    if sheet_name in sheet_names:
        sheet5 = workbook["inventory_sheet"]
    else:
        sheet5 = workbook.create_sheet("inventory_sheet")
    add_inventory = int(input("please tell how many inventory do you want to add"))
    for i in range(add_inventory):
        inventory_id=sheet5.max_row
        inventory_name = input("please tell what inventory do you want to add")
        price = int(input("please tell the price of the inventory"))
        country_orign = input("tell how about the country of origin")
        sheet5.append([inventory_id,inventory_name, price, country_orign])
    workbook.save("pak_nz.xlsx")


def library_management_show():
    workbook = xl.load_workbook("pak_nz.xlsx")
    sheet1 = workbook.active
    for i in sheet1.iter_rows(values_only=True):
        print(i)


def event_management_show():
    workbook = xl.load_workbook("pak_nz.xlsx")
    sheet2 = workbook["event_sheet"]
    for i in sheet2.iter_rows(values_only=True):
        print(i)


def student_management_show():
    workbook = xl.load_workbook("pak_nz.xlsx")
    sheet3 = workbook["student_sheet"]
    for i in sheet3.iter_rows(values_only=True):
        print(i)


def employee_management_show():
    workbook = xl.load_workbook("pak_nz.xlsx")
    sheet4 = workbook["employee_sheet"]
    for i in sheet4.iter_rows(values_only=True):
        print(i)


def inventory_management_show():
    workbook = xl.load_workbook("pak_nz.xlsx")
    sheet5 = workbook["inventory_sheet"]
    for i in sheet5.iter_rows(values_only=True):
        print(i)


def library_management_search():
    workbook = xl.load_workbook("pak_nz.xlsx")
    sheet1 = workbook.active
    search_id = int(input("what id do you want to search for"))
    for i in sheet1.iter_rows(values_only=True):
        if i[0]==search_id:
            print(i)
            break
    else:
        print("no id found")



def event_management_search():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet2=workbook["event_sheet"]
    search_id=int(input("what id do you want to search for"))
    for i in sheet2.iter_rows(values_only=True):
        if i[0]==search_id:
            print(i)
            break
    else:
        print("no id found")

def student_management_search():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet3=workbook["student_sheet"]
    search_id=int(input("what id do you want to to search for"))
    for i in sheet3.iter_rows(values_only=True):
        if i[0]==search_id:
            print(i)
            break
    else:
        print("no id found")

def employee_management_search():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet4=workbook["employee_sheet"]
    search_id=int(input("what id do you want to search for"))
    for i in sheet4.iter_rows(values_only=True):
        if i[0]==search_id:
            print(i)
            break
    else:
        print("no id found")

def inventory_management_search():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet5=workbook["inventory_sheet"]
    search_id=int(input("what id do you want to search for"))
    for i in sheet5.iter_rows(values_only=True):
        if i[0]==search_id:
            print(i)
            break
    else:
        print("no id found")

def library_management_update():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet1=workbook.active
    search_id=int(input("what id do you want to update"))
    row_count=0
    for i in sheet1.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==search_id:
            print("press 1 to update library_book_name")
            print("press 2 to to update author_name")
            print("press 3 to update year_publication")
            choice=int(input("please choose your choice"))
            if choice==1:
                name_book=input("what is the new book name")
                sheet1.cell(row=row_count,column=2,value=name_book)
            elif choice==2:
                author_name=input("what is the new author name")
                sheet1.cell(row=row_count,column=3,value=author_name)
            elif choice==3:
                year_publication=int(input("what is the new year of publication"))
                sheet1.cell(row=row_count,column=4,value=year_publication)
            else:
                print("wrong option chosen")
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def event_management_update():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet2=workbook["event_sheet"]
    search_id=int(input("please choose your id"))
    row_count=0
    for i in sheet2.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==search_id:
            print('press 1 for event_name_update')
            print("press 2 for guests update")
            print("press 3 for expenditure")
            choice=int(input("choose your choice"))
            if choice==1:
                event_name=input("chose your event name")
                sheet2.cell(row=row_count,column=2,value=event_name)
            elif choice==2:
                guests=int(input("what are the new number of guests"))
                sheet2.cell(row=row_count,column=3,value=guests)
            elif choice==3:
                expenditure_update=int(input("what is the new expenditure"))
                sheet2.cell(row=row_count,column=4,value=expenditure_update)
            else:
                print("wrong choice")
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def student_management_update():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet3=workbook["student_sheet"]
    search_id=int(input("which id do you want to search for"))
    row_count=0
    for i in sheet3.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==search_id:
            print("press 1 for student_name_update")
            print("press 2 for grades_update")
            print("press 3 for school_name")
            choice=int(input("choose your choice"))
            if choice==1:
                student_name=input("what is the new student name")
                sheet3.cell(row=row_count,column=2,value=student_name)
            elif choice==2:
                grades=int(input("what are the new grades"))
                sheet3.cell(row=row_count,column=3,value=grades)
            elif choice==3:
                school_update=input("what is the new_school name")
                sheet3.cell(row=row_count,column=4,value=school_update)
            else:
                print("wrong option chosen")
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def employee_management_update():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet4=workbook["employee_sheet"]
    search_id=int(input("what is your search id"))
    row_count=0
    for i in sheet4.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==search_id:
            print("press 1 for employee_name_update")
            print("press 2 for year_experience_update")
            print("press 3 for organization update")
            choice=int(input("choose your choice"))
            if choice==1:
                employee_name=input("choose your new name")
                sheet4.cell(row=row_count,column=2,value=employee_name)
            elif choice==2:
                year_experience_update=int(input("what is new year of experience"))
                sheet4.cell(row=row_count,column=3,value=year_experience_update)
            elif choice==3:
                organization_update=input("what is new organization name")
                sheet4.cell(row=row_count,column=4,value=organization_update)
            else:
                print("wrong choice")
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def inventory_management_update():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet5=workbook["inventory_sheet"]
    search_id=int(input("what is your new id"))
    row_count=0
    for i in sheet5.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==search_id:
            print("press 1 for inventory_name_update")
            print("press 2 for inventory_price")
            print("print 3 for country_origin")
            choice=int(input("choose your choice"))
            if choice==1:
                inventory_name_update=input("please choose a new inventory name")
                sheet5.cell(row=row_count,column=2,value=inventory_name_update)
            elif choice==2:
                inventory_price=input("choose a new inventory price")
                sheet5.cell(row=row_count,column=3,value=inventory_price)
            elif choice==3:
                country_origin_update=input("choose new country update")
                sheet5.cell(row=row_count,column=4,value=country_origin_update)
            else:
                print("wrong choice chosen")
            break

    else:
        print("no search id found")
    workbook.save("pak_nz.xlsx")

def library_management_delete():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet1=workbook["Sheet"]
    delete_id=int(input("what id do you want to delete"))
    row_count=0
    for i in sheet1.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==delete_id:
            sheet1.delete_rows(row_count)
            print("deleted successfully")
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def event_management_delete():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet2=workbook["event_sheet"]
    delete_id=int(input("what id do you want to delete"))
    row_count=0
    for i in sheet2.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==delete_id:
            sheet2.delete_rows(row_count)
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def student_management_delete():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet3=workbook["student_sheet"]
    delete_id=int(input("please select id for deleting"))
    row_count=0
    for i in sheet3.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==delete_id:
            sheet3.delete_rows(row_count)
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def employee_management_delete():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet4=workbook["employee_sheet"]
    delete_id=int(input("please select id for deleting"))
    row_count=0
    for i in sheet4.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==delete_id:
            sheet4.delete_rows(row_count)
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def inventory_management_delete():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet5=workbook["inventory_sheet"]
    delete_id=int(input("please select id"))
    row_count=0
    for i in sheet5.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==delete_id:
            sheet5.delete_rows(row_count)
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")


















