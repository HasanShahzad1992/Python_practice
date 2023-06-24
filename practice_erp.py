import openpyxl as xl
import os

def library_management_add():
    if os.path.exists("pak_nz.xlsx"):
        workbook=xl.load_workbook("pak_nz.xlsx")
    else:
        workbook=xl.Workbook()
    sheetnames=workbook.sheetnames
    sheetname="library_sheet"
    if sheetname in sheetnames:
        sheet1=workbook["library_sheet"]
    else:
        sheet1=workbook.create_sheet("library_sheet")
    add_books=int(input("how many books do you want to enter"))
    for i in range(add_books):
        book_id=sheet1.max_row
        name_book=input("what is the name of the book")
        author_book=input("what is the author name")
        year_publication=int(input("what is the year of publication"))
        sheet1.append([book_id,name_book,author_book,year_publication])
    workbook.save("pak_nz.xlsx")

def event_management_add():
    if os.path.exists("pak_nz.xlsx"):
        workbook=xl.load_workbook("pak_nz.xlsx")
    else:
        workbook=xl.Workbook()
    sheetnames=workbook.sheetnames
    sheetname="event_sheet"
    if sheetname in sheetnames:
        sheet2=workbook["event_sheet"]
    else:
        sheet2=workbook.create_sheet("event_sheet")
    sheet2=workbook["event_sheet"]
    add_events=int(input("how many events do you want to cover"))
    for i in range(add_events):
        event_id=sheet2.max_row
        event_name=input("what is the name of the event")
        guests=int(input("how many guests would be there"))
        expenditure=int(input("what is the expected expenditure"))
        sheet2.append([event_id,event_name,guests,expenditure])
    workbook.save("pak_nz.xlsx")

def student_management_add():
    if os.path.exists("pak_nz.xlsx"):
        workbook=xl.load_workbook("pak_nz.xlsx")
    else:
        workbook=xl.Workbook()
    sheetnames=workbook.sheetnames
    sheetname="student_sheet"
    if sheetname in sheetnames:
        sheet3=workbook["student_sheet"]
    else:
        sheet3=workbook.create_sheet("student_sheet")
    add_students=int(input("how many students do you require"))
    for i in range(add_students):
        student_id=sheet3.max_row
        student_name=input("what is the student name")
        grades=int(input("what is the grade in class"))
        school=input("what is the school name")
        sheet3.append([student_id,student_name,grades,school])
    workbook.save("pak_nz.xlsx")

def library_management_show():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet1=workbook["library_sheet"]
    for i in sheet1.iter_rows(values_only=True):
        print(i)
    workbook.save("pak_nz.xlsx")

def event_management_show():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet2=workbook["event_sheet"]
    for i in sheet2.iter_rows(values_only=True):
        print(i)
    workbook.save("pak_nz.xlsx")

def student_management_show():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet3=workbook["student_sheet"]
    for i in sheet3.iter_rows(values_only=True):
        print(i)
    workbook.save("pak_nz.xlsx")

def library_management_search():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet1=workbook["library_sheet"]
    search_id=int(input("what id do you want to search for"))
    for i in sheet1.iter_rows(values_only=True):
        if i[0]==search_id:
            print(i)
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def event_management_search():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet2=workbook["event_sheet"]
    search_id=int(input("what id do you want to search for"))
    for i in sheet2.iter_rows(values_only=True):
        if i[0]==search_id:
            print(i)
            break
    else:
        print("no id found")

def student_management_search():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet3=workbook["student_sheet"]
    search_id=int(input("what id do you want to search for"))
    for i in sheet3.iter_rows(values_only=True):
        if i[0]==search_id:
            print(i)
            break
    else:
        print("no id found")

def library_management_update():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet1=workbook["library_sheet"]
    update_id=int(input("what id do you want to update"))
    row_count=0
    for i in sheet1.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==update_id:
            print("press 1 for library management update")
            print("press 2 for event management update")
            print("press 3 for student management update")
            choice=int(input("choose  your choice"))
            if choice==1:
                name_book=input("what is the name of the book")
                sheet1.cell(row=row_count,column=2,value=name_book)
            elif choice==2:
                name_author=input("what is the name of the author")
                sheet1.cell(row=row_count,column=3,value=name_author)
            elif choice==3:
                year_publication=int(input("what is the latest year of publication"))
                sheet1.cell(row=row_count,column=4,value=year_publication)
            else:
                print("wrong choice")
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def event_management_update():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet2=workbook["event_sheet"]
    update_id=int(input("what id do you want to update"))
    row_count=0
    for i in sheet2.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==update_id:
            print("press 1 for event_name update")
            print("press 2 for guests update")
            print("press 3 for expenditure update")
            choice=int(input("what is the choice you want to do"))
            if choice==1:
                event_name=input('what is the new event name')
                sheet2.cell(row=row_count,column=2,value=event_name)
            elif choice==2:
                guests_update=int(input("what is the new guest list"))
                sheet2.cell(row=row_count,column=3,value=guests_update)
            elif choice==3:
                expenditure_update=int(input("what is the new expenditure list"))
                sheet2.cell(row=row_count,column=4,value=expenditure_update)
            else:
                print("wrong choice")
            break

    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def student_management_update():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet3=workbook["student_sheet"]
    update_id=int(input("what id do you want to update"))
    row_count=0
    for i in sheet3.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==update_id:
            print("press 1 for student name update")
            print("press 2 for student grades")
            print("press 3 for student school updates")
            choice=int(input('what is your choice'))
            if choice==1:
                student_name=input("what is new student name")
                sheet3.cell(row=row_count,column=2,value=student_name)
            elif choice==2:
                grades=int(input("what is the new grades"))
                sheet3.cell(row=row_count,column=3,value=grades)
            elif choice==3:
                school=input("what is the new school")
                sheet3.cell(row=row_count,column=4,value=school)
            else:
                print("wrong choice")
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def library_management_delete():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet1=workbook["library_sheet"]
    delete_id=int(input("what id do you want to delete"))
    row_count=0
    for i in sheet1.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==delete_id:
            sheet1.delete_rows(row_count)
            break

    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")

def event_management_delete():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet2=workbook["event_sheet"]
    delete_id=int(input("what is the new id to delete"))
    row_count=0
    for i in sheet2.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==delete_id:
            sheet2.delete_rows(row_count)
            break
    else:
        print("no id found")

def student_management_delete():
    workbook=xl.load_workbook("pak_nz.xlsx")
    sheet3=workbook["student_sheet"]
    delete_id=int(input("what is the delete id"))
    row_count=0
    for i in sheet3.iter_rows(values_only=True):
        row_count=row_count+1
        if i[0]==delete_id:
            sheet3.delete_rows(row_count)
            break
    else:
        print("no id found")
    workbook.save("pak_nz.xlsx")



